"""
export_router.py — Excel エクスポートエンドポイント
GET /api/sessions/{sid}/export  → .xlsx ダウンロード

シート構成:
  概要     — 出力日時・データ件数・店舗数など
  分析①〜⑥ — グラフ画像 + 知見・アドバイス + 集計データ表

openpyxl は関数内で遅延インポート（インストール失敗時にアプリ全体がクラッシュするのを防ぐ）
"""
import base64
import io
from datetime import datetime

from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse

import session as sess

router = APIRouter()

_BLUE   = "1E40AF"
_INS_BG = "EFF6FF"   # 知見列 (青系)
_ADV_BG = "F0FDF4"   # アドバイス列 (緑系)
_HDR_BG = "1E40AF"
_HDR_FG = "FFFFFF"
_COL_W  = 56
_IMG_W  = 700
_IMG_H  = 420
_IMG_ROWS = 28


def _build_excel(analyses: list[dict], chat_analyses: list[dict], df, summary_text: str) -> io.BytesIO:
    import openpyxl
    import openpyxl.utils
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Alignment, Font, PatternFill

    def hdr_fill():
        return PatternFill(start_color=_HDR_BG, end_color=_HDR_BG, fill_type="solid")

    def bg_fill(color):
        return PatternFill(start_color=color, end_color=color, fill_type="solid")

    def set_hdr(ws, row, col, value):
        c = ws.cell(row=row, column=col, value=value)
        c.fill = hdr_fill()
        c.font = Font(color=_HDR_FG, bold=True, size=10)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def write_table(ws, table, start_row):
        if not table:
            return
        headers = list(table[0].keys())
        for ci, h in enumerate(headers, 1):
            set_hdr(ws, start_row, ci, h)
            ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = 16
        for ri, row_data in enumerate(table, start=start_row + 1):
            for ci, h in enumerate(headers, 1):
                v = row_data.get(h, "")
                cell = ws.cell(row=ri, column=ci, value=v)
                cell.alignment = Alignment(wrap_text=True)
                if ri % 2 == 0:
                    cell.fill = bg_fill("F8FAFF")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ─── 概要シート ───
    ws0 = wb.create_sheet("概要")
    ws0.column_dimensions["A"].width = 22
    ws0.column_dimensions["B"].width = 44

    ws0["A1"].value = "AIBI4 分析レポート"
    ws0["A1"].font  = Font(bold=True, size=18, color=_BLUE)
    ws0.merge_cells("A1:B1")
    ws0.row_dimensions[1].height = 32

    info = [
        ("出力日時",   datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("データ件数", f"{len(df):,} 行" if df is not None else "—"),
        ("店舗数",     f"{df['店舗名'].nunique()} 店" if df is not None and '店舗名' in df.columns else "—"),
        ("分析件数",   f"{len(analyses)} 項目"),
    ]
    for i, (k, v) in enumerate(info, start=3):
        ws0.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws0.cell(row=i, column=2, value=v)

    if summary_text:
        ws0.cell(row=8, column=1, value="データサマリー").font = Font(bold=True, size=11)
        c = ws0.cell(row=9, column=1, value=summary_text)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws0.merge_cells("A9:B40")
        ws0.row_dimensions[9].height = 400

    # ─── 各分析シート ───
    for a in analyses:
        title = a.get("title", "分析")
        ws = wb.create_sheet(title[:31])
        ws.column_dimensions["A"].width = _COL_W
        ws.column_dimensions["B"].width = _COL_W

        ws["A1"].value = title
        ws["A1"].font  = Font(bold=True, size=13, color=_BLUE)
        ws.merge_cells("A1:B1")
        ws.row_dimensions[1].height = 24

        text_row = 4
        img_b64 = a.get("image_b64")
        if img_b64:
            try:
                xl_img = XLImage(io.BytesIO(base64.b64decode(img_b64)))
                xl_img.width  = _IMG_W
                xl_img.height = _IMG_H
                ws.add_image(xl_img, "A3")
                text_row = 3 + _IMG_ROWS
            except Exception:
                pass

        insights = a.get("insights") or []
        advice   = a.get("advice")   or []

        # 知見ヘッダー
        c = ws.cell(row=text_row, column=1, value="📌 読み取れる知見")
        c.font = Font(bold=True, size=11, color=_BLUE)
        c.fill = bg_fill(_INS_BG)
        c.alignment = Alignment(vertical="center")
        ws.row_dimensions[text_row].height = 20

        c = ws.cell(row=text_row, column=2, value="💼 アドバイス")
        c.font = Font(bold=True, size=11, color="166534")
        c.fill = bg_fill(_ADV_BG)
        c.alignment = Alignment(vertical="center")

        # 知見・アドバイス本文
        n_rows = max(len(insights), len(advice), 1)
        for j in range(n_rows):
            r = text_row + 1 + j
            ws.row_dimensions[r].height = 36
            if j < len(insights):
                c = ws.cell(row=r, column=1, value=f"・{insights[j]}")
                c.fill = bg_fill(_INS_BG)
                c.alignment = Alignment(wrap_text=True, vertical="top")
            if j < len(advice):
                c = ws.cell(row=r, column=2, value=f"・{advice[j]}")
                c.fill = bg_fill(_ADV_BG)
                c.alignment = Alignment(wrap_text=True, vertical="top")

        # 集計データ表
        table = a.get("table")
        if table:
            tbl_start = text_row + n_rows + 2
            ws.cell(row=tbl_start, column=1, value="【集計データ】").font = Font(bold=True, size=11)
            write_table(ws, table, tbl_start + 1)

    # ─── チャット分析シート ───
    if chat_analyses:
        ws_c = wb.create_sheet("チャット分析")
        ws_c.column_dimensions["A"].width = 90

        ws_c["A1"].value = "チャット分析レポート"
        ws_c["A1"].font  = Font(bold=True, size=14, color=_BLUE)
        ws_c.row_dimensions[1].height = 24

        cur = 3
        for idx, entry in enumerate(chat_analyses, 1):
            # 質問行
            q = ws_c.cell(row=cur, column=1, value=f"Q{idx}：{entry.get('question', '')}")
            q.font      = Font(bold=True, size=11, color="FFFFFF")
            q.fill      = bg_fill(_BLUE)
            q.alignment = Alignment(wrap_text=True, vertical="center")
            ws_c.row_dimensions[cur].height = 28
            cur += 1

            # 回答テキスト
            text = entry.get("text", "").strip()
            if text:
                a = ws_c.cell(row=cur, column=1, value=text)
                a.alignment = Alignment(wrap_text=True, vertical="top")
                a.fill      = bg_fill("F4F6FB")
                # 文字数に応じて行高さを調整（上限 200px）
                ws_c.row_dimensions[cur].height = min(max(30, len(text) // 4), 200)
                cur += 1

            # グラフ画像（複数可）
            for g in entry.get("graphs", []):
                img_b64 = g.get("image_b64")
                if not img_b64:
                    continue
                try:
                    xl_img = XLImage(io.BytesIO(base64.b64decode(img_b64)))
                    xl_img.width  = _IMG_W
                    xl_img.height = _IMG_H
                    ws_c.add_image(xl_img, f"A{cur}")
                    cur += _IMG_ROWS + 1
                except Exception:
                    pass

            cur += 2  # エントリー間の余白

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out


@router.get("/sessions/{sid}/export")
def export_excel(sid: str):
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl がインストールされていません。")

    s = sess.get_session(sid)
    if not s:
        raise HTTPException(status_code=404, detail="セッションが見つかりません。")
    analyses = s.get("analyses")
    if not analyses:
        raise HTTPException(status_code=400, detail="分析結果がありません。先に分析を実行してください。")

    df             = s.get("df")
    summary_text   = s.get("summary_text", "")
    chat_analyses  = s.get("chat_analyses") or []

    try:
        buf = _build_excel(analyses, chat_analyses, df, summary_text)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Excel生成エラー: {e}")

    filename = f"AIBI4_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
