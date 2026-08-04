"""分析結果 → Excel レポート（データ＋根拠シート）。

engine.run() の結果 dict を受け取り、
- 「分析」シート: 軸×指標のデータ表
- 「根拠」シート: 使用データ・集計式・軸・適用フィルタ・信頼度・SQL
を持つ .xlsx をバイト列で返す。ThoughtSpot/Copilot 流の「根拠付きレポート化」。
"""
from __future__ import annotations

import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

_HEAD_FILL = PatternFill("solid", fgColor="4F7BD8")
_HEAD_FONT = Font(bold=True, color="FFFFFF")
_KEY_FONT = Font(bold=True, color="4A5163")
_TITLE_FONT = Font(bold=True, size=14, color="1C2333")
_THIN = Side(style="thin", color="D4DAE6")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_DOW = ["月", "火", "水", "木", "金", "土", "日"]
_VIEW_LABEL = {"items": "商品明細（1商品=1行）", "orders": "伝票（1来店=1行）"}

# 指標フォーマット → Excel 表示形式
_NUMFMT = {"yen": "#,##0\"円\"", "int": "#,##0", "float1": "#,##0.0", "pct": "0.0%"}


def _filter_lines(f: dict) -> list[tuple[str, str]]:
    out: list[tuple[str, str]] = []
    if f.get("months"):          out.append(("月", "、".join(map(str, f["months"]))))
    if f.get("stores"):          out.append(("店舗", "、".join(map(str, f["stores"]))))
    if f.get("categories"):      out.append(("カテゴリ", "、".join(map(str, f["categories"]))))
    if f.get("customer_layers"): out.append(("客層", "、".join(map(str, f["customer_layers"]))))
    if f.get("weather"):         out.append(("天気", "、".join(map(str, f["weather"]))))
    if f.get("dow"):             out.append(("曜日", "、".join(_DOW[d] if 0 <= d < 7 else str(d) for d in f["dow"])))
    lo, hi = f.get("hour_min"), f.get("hour_max")
    if lo is not None or hi is not None:
        lo, hi = (lo if lo is not None else 0), (hi if hi is not None else 23)
        out.append(("時間帯", f"{lo}時" if lo == hi else f"{lo}〜{hi}時"))
    return out


def build_report(result: dict, confidence: float | None = None) -> bytes:
    metric = result.get("metric", {})
    dims = result.get("dimensions", [])
    rows = result.get("rows", [])
    spec = result.get("spec", {})
    fmt = _NUMFMT.get(metric.get("fmt", ""), "#,##0")

    wb = Workbook()

    # ── 分析シート ──
    ws = wb.active
    ws.title = "分析"
    ws["A1"] = result.get("title", "分析結果")
    ws["A1"].font = _TITLE_FONT
    ws.append([])

    header = [d.get("label", "項目") for d in dims] + [metric.get("label", "値")]
    hrow = 3
    ws.append(header)
    for ci in range(1, len(header) + 1):
        c = ws.cell(row=hrow, column=ci)
        c.fill, c.font, c.border = _HEAD_FILL, _HEAD_FONT, _BORDER
        c.alignment = Alignment(horizontal="center")

    metric_col = len(header)
    for r in rows:
        line = []
        for i in range(len(dims)):
            line.append(r.get(f"dim{i}"))
        line.append(r.get("value"))
        ws.append(line)
    # データ本体の書式（数値列）
    for ri in range(hrow + 1, hrow + 1 + len(rows)):
        for ci in range(1, len(header) + 1):
            cell = ws.cell(row=ri, column=ci)
            cell.border = _BORDER
            if ci == metric_col:
                cell.number_format = fmt
                cell.alignment = Alignment(horizontal="right")
    if not rows:
        ws.append(["該当データがありません"])

    # 列幅
    for ci in range(1, len(header) + 1):
        width = max(12, *(len(str(ws.cell(row=r, column=ci).value or "")) for r in range(hrow, hrow + len(rows) + 1)))
        ws.column_dimensions[get_column_letter(ci)].width = min(40, width + 4)

    # ── 根拠シート ──
    ev = wb.create_sheet("根拠")
    ev["A1"] = "この分析の根拠"
    ev["A1"].font = _TITLE_FONT
    ev.append([])

    def _kv(key: str, val: str):
        ev.append([key, val])
        r = ev.max_row
        ev.cell(row=r, column=1).font = _KEY_FONT
        ev.cell(row=r, column=1).alignment = Alignment(vertical="top")
        ev.cell(row=r, column=2).alignment = Alignment(vertical="top", wrap_text=True)

    _kv("タイトル", result.get("title", ""))
    _kv("使用データ", _VIEW_LABEL.get(metric.get("view", ""), metric.get("view", "")))
    _kv("指標", f'{metric.get("label", "")}（単位: {metric.get("unit") or "—"}）')
    _kv("集計式", metric.get("agg", ""))
    _kv("軸", " × ".join(d.get("label", "") for d in dims) if dims else "（全体集計）")
    filt = _filter_lines(spec.get("filters", {}) or {})
    _kv("適用フィルタ", "\n".join(f"{k}: {v}" for k, v in filt) if filt else "（なし）")
    if confidence is None:
        confidence = result.get("confidence")
    if confidence is not None:
        _kv("AI信頼度", f"{round(confidence * 100)}%")
    _kv("SQL", result.get("sql", ""))
    _kv("出力日時", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

    ev.column_dimensions["A"].width = 14
    ev.column_dimensions["B"].width = 80

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
