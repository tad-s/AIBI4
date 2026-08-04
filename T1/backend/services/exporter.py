"""T1 PoC Excel出力。"""
from __future__ import annotations

from io import BytesIO

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


def _write_records(ws, title: str, records: list[dict], start_row: int) -> int:
    ws.cell(start_row, 1, title).font = Font(bold=True, size=13)
    row = start_row + 1
    if not records:
        ws.cell(row, 1, "該当なし")
        return row + 2
    headers = list(records[0].keys())
    for col, header in enumerate(headers, 1):
        c = ws.cell(row, col, header)
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="EDEAFF")
    for rec in records:
        row += 1
        for col, header in enumerate(headers, 1):
            ws.cell(row, col, rec.get(header))
    return row + 2


def build_excel(result: dict, funnel: list[dict] | None = None) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "T1 池袋東口店 レコメンドPoC"
    ws["A1"].font = Font(bold=True, size=16)
    summary = result.get("summary", {})
    row = 3
    for key, value in summary.items():
        ws.cell(row, 1, key)
        ws.cell(row, 2, str(value))
        row += 1

    if funnel:
        row += 1
        row = _write_records(ws, "基礎テーブル作成ファネル", funnel, row)

    a1 = result.get("analysis1", {})
    ws1 = wb.create_sheet("1_TOP商品")
    row = _write_records(ws1, "全体TOP10", a1.get("overall", []), 1)
    row = _write_records(ws1, "ドリンクTOP10", a1.get("drink", []), row)
    _write_records(ws1, "フードTOP10", a1.get("food", []), row)

    ws2 = wb.create_sheet("2_同時注文ペア")
    _write_records(ws2, "同時注文ペアTOP10", result.get("analysis2", {}).get("rows", []), 1)

    ws3 = wb.create_sheet("3_連続注文ペア")
    _write_records(ws3, "連続注文ペアTOP10", result.get("analysis3", {}).get("rows", []), 1)

    ws4 = wb.create_sheet("4_推薦候補")
    a4 = result.get("analysis4", {})
    row = _write_records(ws4, "フード→フード Top5", a4.get("food_food", []), 1)
    row = _write_records(ws4, "ドリンク→ドリンク Top5", a4.get("drink_drink", []), row)
    row = _write_records(ws4, "フード↔ドリンク Top5", a4.get("food_drink", []), row)
    _write_records(ws4, "カテゴリ Top5", a4.get("category_top5", []), row)

    for sheet in wb.worksheets:
        for col_cells in sheet.columns:
            max_len = max(len(str(cell.value or "")) for cell in col_cells)
            sheet.column_dimensions[col_cells[0].column_letter].width = min(max(max_len + 2, 12), 48)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
