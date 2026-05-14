# app_v6_9.py  ── Supabase 連携版（居酒屋データ対応）
import io
import json
import time
import re
import calendar
from difflib import get_close_matches
import os
from pathlib import Path
import re
import traceback
from difflib import get_close_matches
from datetime import date
from itertools import combinations
from collections import Counter

import numpy as np
import matplotlib
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import pandas as pd
import streamlit as st
import requests
from dotenv import load_dotenv
from openai import OpenAI
try:
    from streamlit_mic_recorder import mic_recorder
    MIC_AVAILABLE = True
except ImportError:
    MIC_AVAILABLE = False

# ── .env を明示的パスで最優先ロード ──
# load_dotenv() 引数なしは cwd 依存で不安定なため、__file__ 基準の絶対パスを使用
load_dotenv(dotenv_path=Path(__file__).resolve().parent / ".env", override=False)

# ── Streamlit Cloud シークレット → 環境変数へ橋渡し（.env で未設定のキーのみ補完） ──
try:
    import streamlit as _st_secrets
    for _secret_key in ["SUPABASE_URL", "SUPABASE_KEY", "OPENAI_API_KEY"]:
        if _secret_key in _st_secrets.secrets and not os.environ.get(_secret_key):
            os.environ[_secret_key] = _st_secrets.secrets[_secret_key]
except Exception:
    pass

try:
    import openpyxl
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import PatternFill, Font, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# Supabase ローダー
try:
    from supabase_loader import get_client as get_supabase_client
    from supabase_loader import (
        fetch_stores as sb_fetch_stores,
        fetch_available_months,
        fetch_sales_data as sb_fetch_sales_data,
        fetch_visits_for_summary,
        months_to_date_range,
    )
    SUPABASE_AVAILABLE = True
except ImportError:
    SUPABASE_AVAILABLE = False


def count_fetch_chunks(start_date: str, end_date: str) -> int:
    """指定期間のチャンク数を返す（純粋な日付計算・追加クエリなし）。"""
    from datetime import date as _d, timedelta as _td
    CHUNK = 7  # supabase_loader.CHUNK_DAYS と合わせる
    s = _d.fromisoformat(start_date)
    e = _d.fromisoformat(end_date)
    count, cur = 0, s
    while cur <= e:
        cur = min(cur + _td(days=CHUNK - 1), e) + _td(days=1)
        count += 1
    return count

# ── pandas 2.0 互換レイヤー: .append() を pd.concat() で復元 ──
# LLM が生成するコードが古い .append() を使う場合に備えてモンキーパッチ
def _pd_series_append(self, to_append, ignore_index=False, **_kw):
    if not isinstance(to_append, (pd.Series, pd.DataFrame)):
        to_append = pd.Series(to_append)
    return pd.concat([self, to_append], ignore_index=ignore_index)

def _pd_df_append(self, other, ignore_index=False, **_kw):
    if isinstance(other, dict):
        other = pd.DataFrame([other])
    elif isinstance(other, pd.Series):
        other = other.to_frame().T
    return pd.concat([self, other], ignore_index=ignore_index)

if not hasattr(pd.Series, "append") or pd.Series.append is None:
    pd.Series.append = _pd_series_append
if not hasattr(pd.DataFrame, "append") or pd.DataFrame.append is None:
    pd.DataFrame.append = _pd_df_append
# pandas 2.0 では属性自体が存在しないためシンプルに上書き
try:
    pd.Series(dtype=float).append(pd.Series(dtype=float))
except AttributeError:
    pd.Series.append = _pd_series_append
    pd.DataFrame.append = _pd_df_append

# ============================================================
# ===== 追加分析（6項目）共通ユーティリティ =====
# ============================================================

# ── ドリンク・ヘビー・ライト分類キーワード（居酒屋向け） ──
_DRINK_KW = [
    # アルコール類
    "ビール", "生ビール", "生中", "生大", "ハイボール", "チューハイ", "酎ハイ",
    "サワー", "レモンサワー", "梅サワー", "ワイン", "日本酒", "冷酒", "熱燗",
    "焼酎", "麦焼酎", "芋焼酎", "泡盛", "ホッピー", "カクテル", "梅酒",
    "シャンディガフ", "レッドアイ", "シュワ", "スパークリング",
    # ソフトドリンク
    "ウーロン茶", "お茶", "緑茶", "麦茶", "烏龍", "コーラ", "ジュース",
    "ソフトドリンク", "ノンアルコール", "ノンアル", "ドリンク", "ソーダ",
    "オレンジ", "グレープフルーツ", "トマトジュース",
]
_HEAVY_KW = [
    # 揚げ物・ボリューム系
    "唐揚げ", "から揚げ", "フライドチキン", "揚げ", "カツ", "トンカツ",
    "天ぷら", "フライ", "コロッケ", "串カツ", "串揚げ",
    # 焼き物・肉系
    "焼き鳥", "焼鳥", "串焼き", "焼肉", "ステーキ", "ハラミ", "カルビ",
    "豚バラ", "ロース", "ネギ塩", "つくね", "もも",
    # 鍋・煮込み
    "鍋", "おでん", "煮込み", "もつ煮",
    # 炭水化物系
    "ラーメン", "うどん", "そば", "チャーハン", "炒飯", "焼きそば",
    "お茶漬け", "雑炊", "ご飯", "おにぎり",
    # その他ボリューム
    "餃子", "ピザ", "グラタン", "チーズ", "アボカド",
]
_LIGHT_KW = [
    # 野菜・ヘルシー系
    "サラダ", "野菜", "枝豆", "漬物", "キムチ", "冷奴", "豆腐",
    "おひたし", "和え物", "小鉢", "酢の物",
    # 海鮮系（小量）
    "刺身", "刺し身", "お刺身", "カルパッチョ", "マリネ",
    "たこわさ", "いかわさ",
    # 軽い前菜・おつまみ
    "アヒージョ", "ナムル", "ポテサラ", "玉子", "卵焼き",
    "しらす", "おろし",
]


def _kw_match(name, kw_list):
    if pd.isna(name):
        return False
    return any(kw in str(name) for kw in kw_list)


def _build_order_level_df(df: pd.DataFrame) -> pd.DataFrame | None:
    """注文番号単位の DataFrame を構築する（追加分析のベース）。"""
    key_col = next((c for c in ["注文番号", "伝票番号"] if c in df.columns), None)
    if key_col is None:
        return None
    amount_col = next(
        (c for c in ["合計金額(税込)", "合計金額税込"] if c in df.columns), None
    )
    if amount_col is None:
        return None

    d = df.copy()
    d[amount_col] = pd.to_numeric(d[amount_col], errors="coerce")

    if "注文日時" in d.columns:
        d["_dt"] = pd.to_datetime(d["注文日時"], errors="coerce")
    if "希望受取日時" in d.columns:
        d["_recv_dt"] = pd.to_datetime(d["希望受取日時"], errors="coerce")
    # 居酒屋: 来店時間・退店時間
    if "来店時間" in d.columns:
        d["来店時間"] = pd.to_datetime(d["来店時間"], errors="coerce")
    if "退店時間" in d.columns:
        d["退店時間"] = pd.to_datetime(d["退店時間"], errors="coerce")
    if "数量" in d.columns:
        d["数量"] = pd.to_numeric(d["数量"], errors="coerce")
    if "人数" in d.columns:
        d["人数"] = pd.to_numeric(d["人数"], errors="coerce")

    if "商品名" in d.columns:
        d["_is_drink"] = d["商品名"].apply(lambda x: _kw_match(x, _DRINK_KW)).astype(int)
        d["_is_heavy"] = d["商品名"].apply(lambda x: _kw_match(x, _HEAVY_KW)).astype(int)
        d["_is_light"] = d["商品名"].apply(lambda x: _kw_match(x, _LIGHT_KW)).astype(int)

    agg: dict = {"客単価": (amount_col, "first")}
    if "_dt" in d.columns:
        agg["注文日時"] = ("_dt", "first")
    if "_recv_dt" in d.columns:
        agg["受取日時"] = ("_recv_dt", "first")
    if "注文形態" in d.columns:
        agg["注文形態"] = ("注文形態", "first")
    if "店舗名" in d.columns:
        agg["店舗名"] = ("店舗名", "first")
    # 居酒屋: 来店・退店時間
    if "来店時間" in d.columns:
        agg["来店時間"] = ("来店時間", "first")
    if "退店時間" in d.columns:
        agg["退店時間"] = ("退店時間", "first")
    # 居酒屋: 人数・客層
    if "人数" in d.columns:
        agg["人数"] = ("人数", "first")
    if "客層" in d.columns:
        agg["客層"] = ("客層", "first")
    if "商品名" in d.columns:
        agg["商品リスト"] = ("商品名", lambda x: list(x.dropna().astype(str)))
        agg["商品数"] = ("商品名", "count")
    if "数量" in d.columns:
        agg["合計数量"] = ("数量", "sum")
    if "_is_drink" in d.columns:
        agg["ドリンク数"] = ("_is_drink", "sum")
        agg["ヘビー数"] = ("_is_heavy", "sum")
        agg["ライト数"] = ("_is_light", "sum")

    odf = d.groupby(key_col).agg(**agg).reset_index()
    odf.rename(columns={key_col: "注文ID"}, inplace=True)
    odf["客単価"] = pd.to_numeric(odf["客単価"], errors="coerce")
    odf = odf[odf["客単価"] > 0].dropna(subset=["客単価"])

    if "注文日時" in odf.columns:
        odf["曜日"] = odf["注文日時"].dt.dayofweek
        odf["時間帯"] = odf["注文日時"].dt.hour
    elif "来店時間" in odf.columns:
        # 注文日時がなければ来店時間で代替
        odf["曜日"] = odf["来店時間"].dt.dayofweek
        odf["時間帯"] = odf["来店時間"].dt.hour
    if "注文日時" in odf.columns and "受取日時" in odf.columns:
        odf["待ち時間_分"] = (
            (odf["受取日時"] - odf["注文日時"]).dt.total_seconds() / 60
        ).clip(0, 300)
    if "注文形態" in odf.columns:
        odf["店内フラグ"] = odf["注文形態"].str.contains("店内|飲食", na=False).astype(int)
    # 居酒屋: 実際の滞在時間を計算
    if "来店時間" in odf.columns and "退店時間" in odf.columns:
        stay = (odf["退店時間"] - odf["来店時間"]).dt.total_seconds() / 60
        odf["滞在時間_分"] = stay.clip(0, 480)  # 8時間上限
    # 居酒屋: 人数が取れる場合は一人当たり客単価
    if "人数" in odf.columns:
        odf["人数"] = pd.to_numeric(odf["人数"], errors="coerce")
        valid_p = odf["人数"] > 0
        odf.loc[valid_p, "一人単価"] = odf.loc[valid_p, "客単価"] / odf.loc[valid_p, "人数"]
    if "ドリンク数" in odf.columns and "商品数" in odf.columns:
        odf["FD比率"] = (odf["ドリンク数"] / odf["商品数"].replace(0, 1)).clip(0, 1)

    return odf


def _std_regression(X: np.ndarray, y: np.ndarray) -> np.ndarray:
    """標準化して numpy 重回帰 → 標準化係数（β係数）を返す。"""
    xm, xs = X.mean(0), X.std(0)
    xs[xs == 0] = 1.0
    ym, ys = y.mean(), (y.std() or 1.0)
    Xs = (X - xm) / xs
    ys_arr = (y - ym) / ys
    A = np.column_stack([np.ones(len(Xs)), Xs])
    try:
        coef, *_ = np.linalg.lstsq(A, ys_arr, rcond=None)
        return coef[1:]
    except Exception:
        return np.zeros(X.shape[1])


def _fig_to_buf(fig, dpi: int = 150) -> io.BytesIO:
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=dpi, bbox_inches="tight")
    buf.seek(0)
    plt.close(fig)
    return buf


def _placeholder_fig(msg: str) -> io.BytesIO:
    fig, ax = plt.subplots(figsize=(6, 3))
    ax.text(
        0.5, 0.5, msg, ha="center", va="center", fontsize=10, color="#666",
        bbox=dict(boxstyle="round,pad=0.6", facecolor="#f8f9fa", edgecolor="#dee2e6"),
    )
    ax.set_xlim(0, 1)
    ax.set_ylim(0, 1)
    ax.axis("off")
    return _fig_to_buf(fig, dpi=100)


def _dummy_bar(ax, labels, values, title, xlabel="", ylabel="影響度合い", highlight_neg=None):
    """ダミー棒グラフ描画ヘルパー。"""
    colors = []
    for i, (lbl, val) in enumerate(zip(labels, values)):
        if highlight_neg and lbl in highlight_neg:
            colors.append("#e74c3c")
        else:
            colors.append("#5b9bd5")
    bars = ax.bar(labels, [abs(v) for v in values], color=colors, edgecolor="white", linewidth=0.5)
    ax.set_title(title, fontsize=11, pad=8)
    ax.set_ylabel(ylabel)
    if xlabel:
        ax.set_xlabel(xlabel)
    plt.xticks(rotation=30, ha="right", fontsize=8)
    ax.set_ylim(0, max(abs(v) for v in values) * 1.3)


def _insight_advice_block(insight_lines: list[str], advice_lines: list[str]) -> None:
    """📌 知見 + 💼 アドバイス の2列ブロックを表示する。"""
    col_l, col_r = st.columns(2)
    with col_l:
        st.markdown("**📌 読み取れる知見**")
        for line in insight_lines:
            st.markdown(f"- {line}")
    with col_r:
        st.markdown("**💼 アドバイス**")
        for line in advice_lines:
            st.markdown(f"- {line}")


# ============================================================
# ===== 分析① 客単価に影響を与える変数の特定（重回帰分析） =====
# ============================================================

def _analysis_1_variable_regression(order_df: pd.DataFrame | None, return_figs: bool = False):
    """
    return_figs=False: Streamlit に描画（従来動作）
    return_figs=True:  list[tuple[str, Figure, DataFrame|None]] を返す
    """
    figs_out = []

    if not return_figs:
        st.markdown(
            "客単価（伝票1枚あたりの合計金額）を目的変数として、"
            "曜日・時間帯・人数・FD比率・滞在時間などの変数が与える影響度を重回帰分析で可視化します。"
        )
        missing_vars = []
        if order_df is None or "曜日" not in order_df.columns:
            missing_vars.append("注文日時 / 来店時間（曜日・時間帯の抽出に必要）")
        if order_df is None or "人数" not in order_df.columns:
            missing_vars.append("人数（テーブル人数）")
        if order_df is None or "滞在時間_分" not in order_df.columns:
            missing_vars.append("来店時間・退店時間（滞在時間の算出に必要）")
        if missing_vars:
            with st.expander("📋 今回の分析で不足しているデータ"):
                for v in missing_vars:
                    st.markdown(f"- {v}")
                st.caption("※ 不足データは今回の回帰モデルから除外しています。揃えば自動で組み込まれます。")

    use_dummy = True
    coef_df = None

    if order_df is not None and len(order_df) >= 30:
        feature_names = []
        X_list = []
        for fname, col in [
            ("注文時間帯", "時間帯"), ("曜日", "曜日"),
            ("FD比率", "FD比率"), ("商品数", "商品数"),
            ("人数", "人数"), ("滞在時間_分", "滞在時間_分"),
            ("ドリンク数", "ドリンク数"), ("店内フラグ", "店内フラグ"),
        ]:
            if col in order_df.columns:
                s = pd.to_numeric(order_df[col], errors="coerce")
                if s.notna().sum() > 20:
                    feature_names.append(fname)
                    X_list.append(s.values)
        if len(X_list) >= 2:
            tmp = pd.DataFrame(
                dict(zip(feature_names, X_list)),
                index=order_df.index,
            )
            tmp["客単価"] = order_df["客単価"].values
            tmp = tmp.dropna()
            if len(tmp) >= 30:
                X = tmp[feature_names].values.astype(float)
                y = tmp["客単価"].values.astype(float)
                coef = _std_regression(X, y)
                sorted_idx = np.argsort(np.abs(coef))[::-1]
                sorted_names = [feature_names[i] for i in sorted_idx]
                sorted_coef = coef[sorted_idx]
                colors = ["#e74c3c" if c > 0 else "#3498db" for c in sorted_coef]

                fig, ax = plt.subplots(figsize=(8, 4))
                ax.bar(sorted_names, np.abs(sorted_coef), color=colors, edgecolor="white")
                ax.set_ylabel("影響度合い（β係数の絶対値）")
                ax.set_title("客単価への影響度（重回帰分析）")
                legend_els = [
                    mpatches.Patch(color="#e74c3c", label="正の影響（客単価UP方向）"),
                    mpatches.Patch(color="#3498db", label="負の影響（客単価DOWN方向）"),
                ]
                ax.legend(handles=legend_els, fontsize=8)
                plt.xticks(rotation=30, ha="right")
                plt.tight_layout()

                coef_df = pd.DataFrame({"変数名": sorted_names, "β係数": sorted_coef})

                if return_figs:
                    figs_out.append(("客単価への影響度（重回帰分析）", fig, coef_df))
                else:
                    st.image(_fig_to_buf(fig), use_container_width=True)
                    top = sorted_names[0]
                    dir0 = "客単価UP方向" if sorted_coef[0] > 0 else "客単価DOWN方向"
                    insights = [
                        f"**{top}** が客単価に最も強く影響（{dir0}）",
                    ]
                    if len(sorted_names) >= 2:
                        dir1 = "UP方向" if sorted_coef[1] > 0 else "DOWN方向"
                        insights.append(f"2位: **{sorted_names[1]}**（{dir1}）")
                    up_vars = [n for n, c in zip(sorted_names, sorted_coef) if c > 0]
                    dn_vars = [n for n, c in zip(sorted_names, sorted_coef) if c < 0]
                    if up_vars:
                        insights.append(f"客単価を上げる傾向の変数: {', '.join(up_vars[:3])}")
                    if dn_vars:
                        insights.append(f"客単価を下げる傾向の変数: {', '.join(dn_vars[:3])}")
                    advice = []
                    if "FD比率" in up_vars:
                        advice.append("FD比率（フード/ドリンクの比率）が高いほど客単価UP → ドリンク注文を促すオペレーションが有効")
                    if "滞在時間_分" in up_vars:
                        advice.append("滞在時間が長いほど客単価UP → 長時間滞在を促すイベントや居心地改善が売上向上につながる")
                    if "人数" in up_vars:
                        advice.append("来客人数が多いほど客単価UP → グループ来店の促進施策（宴会パック、グループ割引）が有効")
                    if "商品数" in up_vars:
                        advice.append("注文商品数が多いほど客単価UP → 追加注文を促す卓上POPやスタッフ声がけが効果的")
                    if "注文時間帯" in sorted_names[:3]:
                        advice.append("時間帯が客単価に影響 → 時間帯別メニューや限定商品で高単価帯の来客を増やす施策を検討")
                    if not advice:
                        advice.append(f"影響度1位の **{top}** を起点に、客単価向上施策の優先順位を設定する")
                    _insight_advice_block(insights, advice)
                use_dummy = False

    if use_dummy:
        if not return_figs:
            st.warning(
                "⚠️ **ダミーデータによる分析イメージ**  \n"
                "現在のデータで重回帰を実行するには件数か変数が不足しています。"
                "PDFの分析イメージに基づいたダミーデータを表示しています。"
            )
        feat_dummy = ["注文商品", "FD比率", "注文時間帯", "客層", "1組あたり客数", "曜日・時間帯", "その他"]
        val_dummy  = [0.85, 0.72, 0.58, 0.31, 0.28, 0.12, 0.09]
        fig, ax = plt.subplots(figsize=(8, 4))
        _dummy_bar(ax, feat_dummy, val_dummy,
                   title="客単価への影響度（重回帰分析）※ダミーデータ")
        plt.tight_layout()

        coef_df = pd.DataFrame({"変数名": feat_dummy, "β係数（絶対値）": val_dummy})

        if return_figs:
            figs_out.append(("客単価への影響度（重回帰分析）※ダミーデータ", fig, coef_df))
        else:
            st.image(_fig_to_buf(fig), use_container_width=True)
            _insight_advice_block(
                ["注文商品・FD比率・注文時間帯が客単価に大きく影響（ダミー例）"],
                ["ドリンク注文の促進（FD比率向上）や時間帯別施策が客単価UPに有効と想定"],
            )

    if return_figs:
        return figs_out


# ============================================================
# ===== 分析② 客単価に影響を与える商品の特定（重回帰分析） =====
# ============================================================

def _analysis_2_product_regression(df: pd.DataFrame, order_df: pd.DataFrame | None, return_figs: bool = False):
    """
    return_figs=False: Streamlit に描画（従来動作）
    return_figs=True:  list[tuple[str, Figure, DataFrame|None]] を返す
    """
    figs_out = []

    if not return_figs:
        st.markdown(
            "各商品を「注文したか否か（0/1）」の説明変数として重回帰分析を行い、"
            "客単価に対してプラス・マイナスどちらの方向に影響するかを可視化します。"
        )

    use_dummy = True
    coef_df = None

    if (
        order_df is not None
        and "商品リスト" in order_df.columns
        and len(order_df) >= 30
    ):
        try:
            all_items = [item for lst in order_df["商品リスト"] for item in lst if item and item != "nan"]
            freq = Counter(all_items)
            exclude_kw = ["レジ袋", "袋", "クーポン", "割引", "引", "キャンペーン", "0円"]
            top_items = [
                itm for itm, _ in freq.most_common(40)
                if not any(ex in itm for ex in exclude_kw)
            ][:15]

            if len(top_items) >= 5:
                rows = []
                for _, row in order_df.iterrows():
                    item_set = set(row["商品リスト"])
                    enc = {itm: int(itm in item_set) for itm in top_items}
                    enc["客単価"] = row["客単価"]
                    rows.append(enc)
                hot_df = pd.DataFrame(rows).dropna()

                if len(hot_df) >= 30:
                    X = hot_df[top_items].values.astype(float)
                    y = hot_df["客単価"].values.astype(float)
                    coef = _std_regression(X, y)

                    sort_idx = np.argsort(coef)[::-1]
                    top5_pos = [i for i in sort_idx if coef[i] > 0][:5]
                    top5_neg = [i for i in reversed(sort_idx) if coef[i] < 0][:5]
                    show_idx = top5_pos + top5_neg
                    show_names = [top_items[i] for i in show_idx]
                    show_coef  = [coef[i] for i in show_idx]
                    colors = ["#e74c3c" if c > 0 else "#3498db" for c in show_coef]

                    fig, ax = plt.subplots(figsize=(9, 4))
                    ax.bar(show_names, show_coef, color=colors, edgecolor="white")
                    ax.axhline(0, color="black", linewidth=0.8, linestyle="--")
                    ax.set_ylabel("標準化係数（β係数）")
                    ax.set_title("商品別 客単価への影響度（重回帰分析）")
                    legend_els = [
                        mpatches.Patch(color="#e74c3c", label="客単価UP方向"),
                        mpatches.Patch(color="#3498db", label="客単価DOWN方向"),
                    ]
                    ax.legend(handles=legend_els, fontsize=8)
                    plt.xticks(rotation=35, ha="right", fontsize=8)
                    plt.tight_layout()

                    coef_df = pd.DataFrame({"商品名": show_names, "β係数": show_coef})

                    if return_figs:
                        figs_out.append(("商品別 客単価への影響度（重回帰分析）", fig, coef_df))
                    else:
                        st.image(_fig_to_buf(fig), use_container_width=True)
                        pos_items = [show_names[i] for i, c in enumerate(show_coef) if c > 0]
                        neg_items = [show_names[i] for i, c in enumerate(show_coef) if c < 0]
                        insights = []
                        advice = []
                        if pos_items:
                            insights.append(f"**客単価UP商品**: {', '.join(pos_items[:3])} — これらを注文した客ほど伝票単価が高い傾向")
                            advice.append(f"「{pos_items[0]}」などUP商品をおすすめ欄・卓上POPで前面に出し、追加注文を促す")
                        if neg_items:
                            insights.append(f"**客単価DOWN商品**: {', '.join(neg_items[:3])} — 炭水化物系など「締め」商品は満腹感で他の注文を減らす傾向")
                            advice.append(f"「{neg_items[0]}」などDOWN商品は単体推奨より、UP商品とのセット提案で客単価を維持する")
                        if not insights:
                            insights.append("分析結果から有意な傾向が確認されました")
                        if not advice:
                            advice.append("UP商品の推奨強化と、DOWN商品のセット販売を組み合わせた施策を検討")
                        _insight_advice_block(insights, advice)
                    use_dummy = False
        except Exception as e:
            if not return_figs:
                st.warning(f"実データ分析中にエラーが発生しました: {e}")

    if use_dummy:
        if not return_figs:
            st.warning(
                "⚠️ **ダミーデータによる分析イメージ**  \n"
                "実データで分析するには「商品明細（商品名）」と「注文番号」列が必要です。"
            )
        items_d = ["もつ煮込み", "本日のなめろう", "牛たんタタキ", "かんぱち刺し",
                   "やきとん", "牛たん焼きそば"]
        coef_d  = [0.78, 0.65, 0.52, 0.18, 0.10, -0.42]
        colors_d = ["#e74c3c" if c > 0 else "#3498db" for c in coef_d]

        fig, ax = plt.subplots(figsize=(8, 4))
        ax.bar(items_d, coef_d, color=colors_d, edgecolor="white")
        ax.axhline(0, color="black", linewidth=0.8, linestyle="--")
        ax.set_ylabel("標準化係数（β係数）")
        ax.set_title("商品別 客単価への影響度（重回帰分析）※ダミーデータ")
        plt.xticks(rotation=30, ha="right")
        plt.tight_layout()

        coef_df = pd.DataFrame({"商品名": items_d, "β係数": coef_d})

        if return_figs:
            figs_out.append(("商品別 客単価への影響度（重回帰分析）※ダミーデータ", fig, coef_df))
        else:
            st.image(_fig_to_buf(fig), use_container_width=True)
            _insight_advice_block(
                [
                    "客単価UP商品（ダミー例）: もつ煮込み、なめろう、牛たんタタキ — 高単価客ほど頼む傾向",
                    "客単価DOWN商品（ダミー例）: 牛たん焼きそば — 炭水化物系は満腹感で他注文を減らす傾向",
                ],
                [
                    "UP商品をおすすめメニュー欄・卓上POPで強調し、注文率を上げる",
                    "DOWN商品（炭水化物系）はコース終盤に位置づけ、早期注文を避ける声がけが有効",
                ],
            )

    if return_figs:
        return figs_out


# ============================================================
# ===== 分析③ ABC分析 + グループ別商品分析 =====
# ============================================================

def _analysis_3_abc_analysis(df: pd.DataFrame, order_df: pd.DataFrame | None, return_figs: bool = False):
    """
    return_figs=False: Streamlit に描画（従来動作）
    return_figs=True:  list[tuple[str, Figure, DataFrame|None]] を返す
    """
    figs_out = []

    if not return_figs:
        st.markdown(
            "顧客（注文）を客単価の高・中・低でグルーピング（ABC分析）し、"
            "各グループで売上に貢献している商品と、炭水化物系商品の注文タイミングを比較します。"
        )

    use_dummy_timing = True

    if order_df is not None and len(order_df) >= 30 and "商品リスト" in order_df.columns:
        try:
            odf = order_df.copy()
            q33 = odf["客単価"].quantile(0.33)
            q67 = odf["客単価"].quantile(0.67)

            def _grade(v):
                if v >= q67:
                    return "高（A）"
                elif v >= q33:
                    return "中（B）"
                else:
                    return "低（C）"

            odf["客単価グループ"] = odf["客単価"].apply(_grade)
            group_stats = odf.groupby("客単価グループ").agg(
                件数=("客単価", "count"),
                平均客単価=("客単価", "mean"),
            ).reindex(["高（A）", "中（B）", "低（C）"])
            avg_all = odf["客単価"].mean()

            # ── グループ分布グラフ ──
            fig1, ax1 = plt.subplots(figsize=(5, 3.5))
            bar_colors = ["#c0392b", "#e67e22", "#27ae60"]
            ax1.bar(group_stats.index, group_stats["平均客単価"],
                   color=bar_colors, edgecolor="white")
            ax1.axhline(avg_all, color="orange", linewidth=2,
                       linestyle="--", label=f"全体平均 {avg_all:,.0f}円")
            ax1.set_ylabel("平均客単価（円）")
            ax1.set_title("客単価グループ別 平均客単価（ABC分析）")
            ax1.legend(fontsize=8)
            plt.tight_layout()

            if return_figs:
                figs_out.append(("ABC分析 - グループ別平均客単価", fig1, group_stats.reset_index()))
            else:
                col_l, col_r = st.columns(2)
                with col_l:
                    st.markdown("**客単価グループ分布**")
                    st.image(_fig_to_buf(fig1), use_container_width=True)

            # ── グループ別トップ商品グラフ ──
            exclude_kw = ["レジ袋", "袋", "クーポン", "割引", "引", "0円"]
            fig2, axes = plt.subplots(1, 3, figsize=(9, 3.5))
            for idx, (grp, clr) in enumerate([("高（A）", "#c0392b"), ("中（B）", "#e67e22"), ("低（C）", "#27ae60")]):
                sub = odf[odf["客単価グループ"] == grp]
                items = [
                    itm for lst in sub["商品リスト"] for itm in lst
                    if itm and not any(ex in itm for ex in exclude_kw)
                ]
                top5 = [itm for itm, _ in Counter(items).most_common(5)]
                cnts = [Counter(items)[itm] for itm in top5]
                axes[idx].barh(top5[::-1], cnts[::-1], color=clr, edgecolor="white")
                axes[idx].set_title(f"{grp}", fontsize=9)
                axes[idx].tick_params(axis="y", labelsize=7)
            plt.suptitle("グループ別 売上貢献商品 Top5", fontsize=10)
            plt.tight_layout()

            if return_figs:
                figs_out.append(("ABC分析 - グループ別売上貢献商品", fig2, None))
            else:
                with col_r:
                    st.markdown("**グループ別 売上貢献商品 Top5**")
                    st.image(_fig_to_buf(fig2), use_container_width=True)
                a_avg = group_stats.loc["高（A）", "平均客単価"]
                c_avg = group_stats.loc["低（C）", "平均客単価"]
                gap = a_avg - c_avg
                a_cnt = int(group_stats.loc["高（A）", "件数"])
                c_cnt = int(group_stats.loc["低（C）", "件数"])
                _insight_advice_block(
                    [
                        f"全体平均客単価: **{avg_all:,.0f}円**",
                        f"高単価グループ（A）平均: {a_avg:,.0f}円（{a_cnt}件） / 低単価グループ（C）平均: {c_avg:,.0f}円（{c_cnt}件）",
                        f"A・C間の客単価差は **{gap:,.0f}円** — 上位客の購買行動が売上を大きく左右している",
                    ],
                    [
                        f"低単価グループ（C）客に追加注文を促す施策（{c_avg:,.0f}円 → 中間帯への引き上げ）が全体売上向上に直結",
                        "Cグループに人気のサイドメニューやデザートを低価格帯で提案し、購入点数を増やす",
                        "Aグループのリピート来店を促進（ポイントカード、優待案内）し高単価客の離脱を防ぐ",
                    ],
                )

            use_dummy_timing = False

        except Exception as e:
            if not return_figs:
                st.warning(f"ABC分析中にエラーが発生しました: {e}")

    if use_dummy_timing:
        if not return_figs:
            st.warning(
                "⚠️ **ダミーデータによる分析イメージ**  \n"
                "実データ分析には「注文番号」「合計金額(税込)」「商品名」が必要です。"
            )
        grps = ["高（A）", "中（B）", "低（C）"]
        avgs = [7200, 4500, 2100]
        fig_d, ax_d = plt.subplots(figsize=(5, 3.5))
        ax_d.bar(grps, avgs, color=["#c0392b", "#e67e22", "#27ae60"], edgecolor="white")
        ax_d.axhline(3000, color="orange", linewidth=2, linestyle="--", label="平均 3,000円")
        ax_d.set_ylabel("平均客単価（円）")
        ax_d.set_title("客単価グループ別 平均客単価（ABC分析）※ダミーデータ")
        ax_d.legend(fontsize=8)
        plt.tight_layout()
        dummy_df = pd.DataFrame({"グループ": grps, "平均客単価": avgs})
        if return_figs:
            figs_out.append(("ABC分析 - グループ別平均客単価（ダミーデータ）", fig_d, dummy_df))
        else:
            st.image(_fig_to_buf(fig_d), use_container_width=True)

    # ── 注文タイミング（ダミー）──
    if not return_figs:
        st.markdown("**炭水化物系メニューの注文タイミング比較（グループ別）**")
        st.info(
            "📋 注文タイミングの分析には各商品の個別注文時刻が必要ですが、"
            "現在のデータでは伝票単位で1つの注文日時しか記録されていません。"
            "以下はダミーデータによる分析イメージです。"
        )
        with st.expander("必要な追加データ"):
            st.markdown(
                "- **個別商品の注文時刻**（注文明細ごとの時刻記録）\n"
                "- **初回注文からの経過時間**（FD比率の元データ）"
            )

    time_slots = ["1", "2", "3", "4", "5", "6"]
    high_pct  = [7, 5, 10, 12, 28, 20]
    mid_pct   = [10, 19, 30, 30, 10, 5]
    low_pct   = [45, 20, 20, 10, 10, 5]
    x = np.arange(len(time_slots))
    w = 0.25
    fig_timing, ax_t = plt.subplots(figsize=(8, 4))
    ax_t.bar(x - w, high_pct, w, label="高客単価グループ（A）", color="#c0392b", edgecolor="white")
    ax_t.bar(x,     mid_pct,  w, label="中客単価グループ（B）", color="#e67e22", edgecolor="white")
    ax_t.bar(x + w, low_pct,  w, label="低客単価グループ（C）", color="#27ae60", edgecolor="white")
    ax_t.set_xlabel("注文タイミング（初回注文からの経過ブロック）")
    ax_t.set_ylabel("注文割合（%）")
    ax_t.set_title("炭水化物系メニューの注文タイミング比較 ※ダミーデータ")
    ax_t.set_xticks(x)
    ax_t.set_xticklabels(time_slots)
    ax_t.legend(fontsize=8)
    plt.tight_layout()

    if return_figs:
        figs_out.append(("ABC分析 - 炭水化物注文タイミング（ダミー）", fig_timing, None))
    else:
        st.image(_fig_to_buf(fig_timing), use_container_width=True)
        st.info(
            "💡 ダミー例：高単価グループは炭水化物系を「締め」として遅いタイミングで注文する傾向。"
            "低単価グループは最初から注文してお腹を満たしてしまう傾向が示唆されます。"
        )

    if return_figs:
        return figs_out


# ============================================================
# ===== 分析④ マーケットバスケット分析（一緒に注文されるメニュー） =====
# ============================================================

_DRINK_KEYWORDS = [
    "コーヒー", "ラテ", "エスプレッソ", "カプチーノ", "ティー", "紅茶", "緑茶",
    "ジュース", "スムージー", "フラペ", "アメリカーノ", "マキアート", "モカ",
    "カフェオレ", "ソーダ", "レモネード", "ミルク", "ホットチョコ", "ドリンク",
    "ビール", "サワー", "ハイボール", "日本酒", "ワイン", "チューハイ",
    "ウーロン", "ウイスキー", "焼酎", "梅酒", "酎ハイ", "ソフトドリンク",
    "烏龍", "コーラ", "ジンジャー",
]

def _item_category(name: str) -> str:
    return "ドリンク" if any(kw in name for kw in _DRINK_KEYWORDS) else "フード"


def _analysis_4_basket_analysis(df: pd.DataFrame, order_df: pd.DataFrame | None, return_figs: bool = False):
    """
    return_figs=False: Streamlit に描画（従来動作）
    return_figs=True:  list[tuple[str, Figure, DataFrame|None]] を返す
    """
    figs_out = []

    if not return_figs:
        st.markdown(
            "同一注文内で一緒に頼まれる商品の組み合わせを集計します。"
            "全体の共起傾向と、ドリンク×フードのクロスカテゴリ組み合わせを合わせて表示します。"
        )

    use_dummy = True

    if order_df is not None and "商品リスト" in order_df.columns and len(order_df) >= 20:
        try:
            exclude_kw = ["レジ袋", "袋", "クーポン", "割引", "引", "0円"]

            def clean_list(lst):
                return [
                    itm for itm in lst
                    if itm and not any(ex in itm for ex in exclude_kw)
                ]

            clean_orders = [clean_list(lst) for lst in order_df["商品リスト"]]
            clean_orders = [lst for lst in clean_orders if len(lst) >= 2]

            if len(clean_orders) >= 10:
                all_items = [itm for lst in clean_orders for itm in lst]
                freq = Counter(all_items)
                top_items = [itm for itm, _ in freq.most_common(12)]

                pair_count: Counter = Counter()
                for lst in clean_orders:
                    filtered = [itm for itm in lst if itm in top_items]
                    for a, b in combinations(set(filtered), 2):
                        pair_count[tuple(sorted([a, b]))] += 1

                # ─── ① 全体の共起傾向 ───
                n = len(top_items)
                mat = np.zeros((n, n))
                item_idx = {itm: i for i, itm in enumerate(top_items)}
                for (a, b), cnt in pair_count.items():
                    if a in item_idx and b in item_idx:
                        i, j = item_idx[a], item_idx[b]
                        mat[i, j] = cnt
                        mat[j, i] = cnt

                fig1, ax1 = plt.subplots(figsize=(7, 6))
                im = ax1.imshow(mat, cmap="YlOrRd", aspect="auto")
                ax1.set_xticks(range(n))
                ax1.set_yticks(range(n))
                short = [itm[:10] for itm in top_items]
                ax1.set_xticklabels(short, rotation=45, ha="right", fontsize=7)
                ax1.set_yticklabels(short, fontsize=7)
                plt.colorbar(im, ax=ax1, shrink=0.8)
                ax1.set_title("商品共起頻度ヒートマップ（全体）")
                plt.tight_layout()

                top_pairs = pair_count.most_common(10)
                pair_df = pd.DataFrame(
                    [(f"{a} × {b}", cnt) for (a, b), cnt in top_pairs],
                    columns=["商品ペア", "共起件数"],
                )

                # ─── ② ドリンク×フード クロスカテゴリ ───
                cross_counter = Counter({
                    pair: cnt for pair, cnt in pair_count.items()
                    if _item_category(pair[0]) != _item_category(pair[1])
                })
                top_cross = cross_counter.most_common(10)
                cross_df = pd.DataFrame(
                    [(f"{a} × {b}", cnt, f"{_item_category(a)} × {_item_category(b)}")
                     for (a, b), cnt in top_cross],
                    columns=["商品ペア", "共起件数", "カテゴリ"],
                )

                fig2 = None
                if top_cross:
                    labels = [f"{a} × {b}" for (a, b), _ in top_cross]
                    vals   = [cnt for _, cnt in top_cross]
                    fig2, ax2 = plt.subplots(figsize=(7, max(4, len(labels) * 0.45)))
                    bars = ax2.barh(labels[::-1], vals[::-1], color="#5b9bd5", edgecolor="white")
                    for bar, v in zip(bars, vals[::-1]):
                        ax2.text(bar.get_width() + max(vals) * 0.01,
                                 bar.get_y() + bar.get_height() / 2,
                                 f"{v}件", va="center", fontsize=8)
                    ax2.set_xlabel("共起件数")
                    ax2.set_title("ドリンク × フード クロスカテゴリ Top10")
                    ax2.set_xlim(0, max(vals) * 1.2)
                    plt.tight_layout()

                if return_figs:
                    figs_out.append(("バスケット① 共起頻度ヒートマップ（全体）", fig1, pair_df))
                    if fig2 is not None:
                        figs_out.append(("バスケット② ドリンク×フード クロスカテゴリ", fig2, cross_df))
                else:
                    st.markdown("#### ① 全体の共起傾向")
                    col_l, col_r = st.columns([3, 2])
                    with col_l:
                        st.markdown("**共起頻度ヒートマップ（上位12商品）**")
                        st.image(_fig_to_buf(fig1), use_container_width=True)
                    with col_r:
                        st.markdown("**共起ランキング Top10**")
                        st.dataframe(pair_df, use_container_width=True, hide_index=True)
                    st.info(
                        "💡 **全体の傾向**: グループ注文ではテーブル全員がそれぞれドリンクを注文するため、"
                        "ドリンク同士の組み合わせが上位に入りやすい傾向があります。"
                        "テーブル単位の注文構造や人気商品の共起パターンを把握するのに有効です。"
                    )

                    st.markdown("#### ② ドリンク × フード クロスカテゴリ組み合わせ")
                    if top_cross:
                        col_l2, col_r2 = st.columns([3, 2])
                        with col_l2:
                            st.image(_fig_to_buf(fig2), use_container_width=True)
                        with col_r2:
                            st.dataframe(cross_df[["商品ペア", "共起件数"]], use_container_width=True, hide_index=True)
                        best = top_cross[0]
                        top3_cross = [f"{a} × {b}" for (a, b), _ in top_cross[:3]]
                        _insight_advice_block(
                            [
                                f"クロスカテゴリ No.1: **{best[0][0]}** × **{best[0][1]}**（{best[1]}件）",
                                f"上位3ペア: {', '.join(top3_cross)} — これらがセット注文の核心",
                                "ドリンク×フードの組み合わせは、テーブル全体の注文量拡大に直結するシグナル",
                            ],
                            [
                                f"「{best[0][0]}」×「{best[0][1]}」をセットメニュー化し、単品注文より割安感を演出する",
                                "上位クロスペアを卓上POPや口頭推奨に活用し、追加注文率を高める",
                                "フードを注文した客にドリンクを勧める（またはその逆）スクリプトをスタッフに共有",
                            ],
                        )
                    else:
                        st.info("ドリンク×フードのクロスカテゴリペアが見つかりませんでした。")

                use_dummy = False
        except Exception as e:
            if not return_figs:
                st.warning(f"バスケット分析中にエラーが発生しました: {e}")

    if use_dummy:
        if not return_figs:
            st.warning(
                "⚠️ **ダミーデータによる分析イメージ**  \n"
                "実データで分析するには「注文番号」「商品名」列が必要で、"
                "2品以上の注文が10件以上あることが条件です。"
            )
        items_d = ["バーガー", "ポテト", "ドリンク", "ナゲット", "サラダ",
                   "アップル", "コーヒー", "チキン"]
        n_d = len(items_d)
        mat_d = np.array([
            [0,  85, 90, 40, 20, 15, 30, 35],
            [85,  0, 78, 55, 18, 12, 20, 40],
            [90, 78,  0, 38, 25, 20, 60, 33],
            [40, 55, 38,  0, 10,  8, 15, 28],
            [20, 18, 25, 10,  0, 30, 22, 12],
            [15, 12, 20,  8, 30,  0, 18,  9],
            [30, 20, 60, 15, 22, 18,  0, 14],
            [35, 40, 33, 28, 12,  9, 14,  0],
        ], dtype=float)
        fig_d, ax_d = plt.subplots(figsize=(6, 5))
        im_d = ax_d.imshow(mat_d, cmap="YlOrRd", aspect="auto")
        ax_d.set_xticks(range(n_d))
        ax_d.set_yticks(range(n_d))
        ax_d.set_xticklabels(items_d, rotation=45, ha="right", fontsize=8)
        ax_d.set_yticklabels(items_d, fontsize=8)
        plt.colorbar(im_d, ax=ax_d, shrink=0.8)
        ax_d.set_title("商品共起頻度ヒートマップ ※ダミーデータ")
        plt.tight_layout()

        dummy_pairs = pd.DataFrame({
            "商品ペア": ["バーガー × ドリンク", "バーガー × ポテト", "ドリンク × コーヒー"],
            "共起件数": [90, 85, 60],
        })

        if return_figs:
            figs_out.append(("バスケット分析 - 共起頻度ヒートマップ（ダミー）", fig_d, dummy_pairs))
        else:
            st.image(_fig_to_buf(fig_d), use_container_width=True)
            st.info("💡 ダミー例：バーガー×ドリンク・バーガー×ポテトの組み合わせが最も多い。")

    if return_figs:
        return figs_out


# ============================================================
# ===== 分析⑤ ヘビー系 vs ライト系 食べ物と客単価の傾向 =====
# ============================================================

def _analysis_5_heavy_light(df: pd.DataFrame, order_df: pd.DataFrame | None, return_figs: bool = False):
    """
    return_figs=False: Streamlit に描画（従来動作）
    return_figs=True:  list[tuple[str, Figure, DataFrame|None]] を返す
    """
    figs_out = []

    if not return_figs:
        st.markdown(
            "注文商品のキーワードから「ヘビー系（揚げ物・大きいサイズなど）」と"
            "「ライト系（サラダ・小さいサイズなど）」に分類し、"
            "グループ別の客単価傾向と注文点数を比較します。"
        )

    use_dummy = True

    if (
        order_df is not None
        and "ヘビー数" in order_df.columns
        and "ライト数" in order_df.columns
        and len(order_df) >= 20
    ):
        try:
            odf = order_df.copy()

            def classify(row):
                h, l = row.get("ヘビー数", 0), row.get("ライト数", 0)
                if h == 0 and l == 0:
                    return "その他"
                if h > l:
                    return "ヘビー系"
                if l > h:
                    return "ライト系"
                return "ミックス"

            odf["食品タイプ"] = odf.apply(classify, axis=1)
            grp = (
                odf.groupby("食品タイプ")["客単価"]
                .agg(["mean", "median", "count"])
                .rename(columns={"mean": "平均客単価", "median": "中央値", "count": "件数"})
                .loc[lambda d: d["件数"] >= 5]
            )

            if len(grp) >= 2:
                clr_map = {
                    "ヘビー系": "#e74c3c", "ライト系": "#27ae60",
                    "ミックス": "#e67e22", "その他": "#95a5a6",
                }
                colors = [clr_map.get(g, "#999") for g in grp.index]

                fig, ax = plt.subplots(figsize=(6, 4))
                ax.bar(grp.index, grp["平均客単価"], color=colors, edgecolor="white")
                ax.set_ylabel("平均客単価（円）")
                ax.set_title("食品タイプ別 平均客単価")
                plt.tight_layout()

                if return_figs:
                    figs_out.append(("ヘビー/ライト分析 - 食品タイプ別平均客単価", fig, grp.reset_index()))
                else:
                    col_l, col_r = st.columns(2)
                    with col_l:
                        st.markdown("**食品タイプ別 平均客単価**")
                        st.image(_fig_to_buf(fig), use_container_width=True)
                    with col_r:
                        st.markdown("**食品タイプ別 件数・客単価サマリ**")
                        st.dataframe(grp.style.format({"平均客単価": "{:,.0f}円", "中央値": "{:,.0f}円"}),
                                     use_container_width=True)
                    if "ヘビー系" in grp.index and "ライト系" in grp.index:
                        h_avg = grp.loc["ヘビー系", "平均客単価"]
                        l_avg = grp.loc["ライト系", "平均客単価"]
                        h_cnt = int(grp.loc["ヘビー系", "件数"])
                        l_cnt = int(grp.loc["ライト系", "件数"])
                        if h_avg > l_avg:
                            diff = h_avg - l_avg
                            _insight_advice_block(
                                [
                                    f"ヘビー系客の平均客単価（{h_avg:,.0f}円 / {h_cnt}件）はライト系（{l_avg:,.0f}円 / {l_cnt}件）より **{diff:,.0f}円高い**",
                                    "揚げ物・大きいサイズを選ぶ客ほど全体的に注文量が多い傾向",
                                    "ヘビー系客はコース外での追加注文も積極的に行う可能性が高い",
                                ],
                                [
                                    "ヘビー系商品（揚げ物・大サイズ）を注文した客に追加ドリンクや締めメニューを積極的に提案",
                                    "大盛り・サイズアップオプションを設けてアップセルを促進",
                                    "ヘビー系客の来店率を高める「がっつり系」コースやランチ限定大盛りメニューを検討",
                                ],
                            )
                        else:
                            diff = l_avg - h_avg
                            _insight_advice_block(
                                [
                                    f"ライト系客の平均客単価（{l_avg:,.0f}円 / {l_cnt}件）はヘビー系（{h_avg:,.0f}円 / {h_cnt}件）より **{diff:,.0f}円高い**",
                                    "ライト系客は1品あたりの量より品数を多く注文する傾向 — 薄利多売型の購買パターン",
                                    "サラダ・小皿系メニューを起点に多品注文が発生している可能性",
                                ],
                                [
                                    "ライト系メニューのバリエーションを増やし、多品注文を誘発する品揃えにする",
                                    "小皿・シェアメニューをグループ向けにセット提案し、品数増加を促す",
                                    "ライト系客向けに「おまかせ小皿コース」など多品構成のコースを用意する",
                                ],
                            )
                use_dummy = False
        except Exception as e:
            if not return_figs:
                st.warning(f"分析中にエラーが発生しました: {e}")

    if use_dummy:
        if not return_figs:
            st.warning(
                "⚠️ **ダミーデータによる分析イメージ**  \n"
                "実データ分析には「商品名」列が必要です（キーワード分類で自動判定します）。"
            )
        types_d = ["ヘビー系\n（揚げ物・大きいサイズ）", "ライト系\n（サラダ・小さいサイズ）",
                   "ミックス", "その他"]
        avg_d    = [3800, 2900, 3200, 2500]
        cnt_d    = [320, 180, 250, 90]
        colors_d = ["#e74c3c", "#27ae60", "#e67e22", "#95a5a6"]

        fig_d, (ax1, ax2) = plt.subplots(1, 2, figsize=(9, 4))
        ax1.bar(types_d, avg_d, color=colors_d, edgecolor="white")
        ax1.set_ylabel("平均客単価（円）")
        ax1.set_title("食品タイプ別 平均客単価 ※ダミーデータ")
        ax2.bar(types_d, cnt_d, color=colors_d, edgecolor="white")
        ax2.set_ylabel("注文件数")
        ax2.set_title("食品タイプ別 注文件数 ※ダミーデータ")
        for ax in [ax1, ax2]:
            plt.setp(ax.get_xticklabels(), fontsize=8)
        plt.tight_layout()

        dummy_df = pd.DataFrame({
            "食品タイプ": ["ヘビー系", "ライト系", "ミックス", "その他"],
            "平均客単価": avg_d,
            "注文件数": cnt_d,
        })

        if return_figs:
            figs_out.append(("ヘビー/ライト分析（ダミーデータ）", fig_d, dummy_df))
        else:
            st.image(_fig_to_buf(fig_d), use_container_width=True)
            _insight_advice_block(
                [
                    "ヘビー系（揚げ物・大サイズ）客の客単価が最も高い傾向（ダミー例）",
                    "ライト系客も品数が多く、一定の客単価を維持している",
                ],
                [
                    "大きいサイズへのアップグレード誘導で客単価向上が見込める（ダミー想定）",
                    "ライト系客には多品注文を促す小皿セットやシェアプレートを提案",
                ],
            )

    if return_figs:
        return figs_out


# ============================================================
# ===== 分析⑥ 滞留時間別客単価（時間客単価算出） =====
# ============================================================

def _analysis_6_stay_time_unit_price(df: pd.DataFrame, order_df: pd.DataFrame | None, return_figs: bool = False):
    """
    居酒屋版: visits.visit_time / leave_time から取得した実際の滞在時間を使用。
    return_figs=False: Streamlit に描画（従来動作）
    return_figs=True:  list[tuple[str, Figure, DataFrame|None]] を返す
    """
    figs_out = []

    if not return_figs:
        st.markdown(
            "来店時間・退店時間から算出した**実際の滞在時間**別に客単価を集計し、"
            "「時間単位の売上効率（時間客単価）」を算出します。"
        )

    # ── 実際の滞在時間データが存在する場合（居酒屋: 通常ここに入る） ──
    has_real_stay = (
        order_df is not None
        and "滞在時間_分" in order_df.columns
        and order_df["滞在時間_分"].notna().sum() >= 20
    )

    if has_real_stay:
        stay_df = order_df[
            (order_df["滞在時間_分"] > 0) & (order_df["滞在時間_分"] <= 360)
        ].copy()

        if len(stay_df) >= 20:
            bins   = [0, 30, 60, 90, 120, 150, 360]
            labels = ["0-30分", "30-60分", "60-90分", "90-120分", "120-150分", "150分超"]
            stay_df["滞在時間帯"] = pd.cut(stay_df["滞在時間_分"], bins=bins, labels=labels)
            grp = stay_df.groupby("滞在時間帯", observed=True)["客単価"].agg(
                ["mean", "count"]
            ).rename(columns={"mean": "平均客単価", "count": "件数"})
            grp = grp[grp["件数"] >= 3]

            if len(grp) >= 2:
                # 時間客単価 = 客単価 / (滞在時間/60)
                stay_center = {"0-30分": 0.25, "30-60分": 0.75, "60-90分": 1.25,
                               "90-120分": 1.75, "120-150分": 2.25, "150分超": 3.0}
                # CategoricalIndex.map は全カテゴリに適用されるため iterrows で計算
                grp["時間客単価"] = [
                    row["平均客単価"] / stay_center.get(str(idx), 1.0)
                    if stay_center.get(str(idx), 0) > 0 else 0
                    for idx, row in grp.iterrows()
                ]

                fig_real, (ax1, ax2) = plt.subplots(1, 2, figsize=(12, 4))

                ax1.bar(grp.index.astype(str), grp["平均客単価"],
                        color="#5b9bd5", edgecolor="white")
                ax1.set_xlabel("滞在時間帯")
                ax1.set_ylabel("平均客単価（円）")
                ax1.set_title("滞在時間帯別 平均客単価")
                for i, (v, c) in enumerate(zip(grp["平均客単価"], grp["件数"])):
                    ax1.text(i, v + 50, f"{int(v):,}円\n(n={c})", ha="center", fontsize=8)
                ax1.tick_params(axis="x", rotation=20)

                ax2.plot(range(len(grp)), grp["時間客単価"].values,
                         marker="o", color="#e74c3c", linewidth=2)
                ax2.fill_between(range(len(grp)), grp["時間客単価"].values,
                                 alpha=0.2, color="#e74c3c")
                ax2.set_xticks(range(len(grp)))
                ax2.set_xticklabels(grp.index.astype(str), rotation=20, ha="right")
                ax2.set_ylabel("時間客単価（円/時間）")
                ax2.set_title("滞在時間帯別 時間客単価")

                plt.tight_layout()

                if return_figs:
                    figs_out.append(("滞在時間分析 - 客単価・時間客単価",
                                     fig_real, grp.reset_index()))
                else:
                    st.image(_fig_to_buf(fig_real), use_container_width=True)
                    max_usp_idx = grp["時間客単価"].idxmax()
                    max_usp_val = grp.loc[max_usp_idx, "時間客単価"]
                    max_sp_idx = grp["平均客単価"].idxmax()
                    max_sp_val = grp.loc[max_sp_idx, "平均客単価"]
                    _insight_advice_block(
                        [
                            f"時間あたり売上効率（時間客単価）が最も高い滞在帯: **{max_usp_idx}**（{max_usp_val:,.0f}円/時）",
                            f"客単価の絶対値が最も高い滞在帯: **{max_sp_idx}**（{max_sp_val:,.0f}円）— 長時間ほど伝票単価は大きい",
                            "短時間帯は時間客単価が高く『効率型』、長時間帯は客単価が高く『高単価型』の2パターンが存在",
                        ],
                        [
                            f"ピーク時は{max_usp_idx}の回転数を最大化し、時間当たり売上を向上させる",
                            "時間制コース（例: 90分制）の価格を時間客単価に基づいて最適化する",
                            "閑散時は長時間滞在客の追加注文を促す施策（デザート・追加ドリンク声がけ）で高単価化を図る",
                        ],
                    )

                # 人数×滞在時間の散布図（人数データがある場合）
                if "人数" in stay_df.columns and stay_df["人数"].notna().sum() >= 10:
                    fig_scatter, ax_s = plt.subplots(figsize=(7, 4))
                    sc = ax_s.scatter(
                        stay_df["滞在時間_分"], stay_df["客単価"],
                        c=pd.to_numeric(stay_df["人数"], errors="coerce"),
                        cmap="YlOrRd", alpha=0.6, edgecolors="white", linewidth=0.5
                    )
                    plt.colorbar(sc, ax=ax_s, label="人数")
                    ax_s.set_xlabel("滞在時間（分）")
                    ax_s.set_ylabel("客単価（円）")
                    ax_s.set_title("滞在時間 × 客単価（人数で色分け）")
                    plt.tight_layout()
                    if return_figs:
                        figs_out.append(("滞在時間×客単価散布図（人数色分け）", fig_scatter, None))
                    else:
                        st.image(_fig_to_buf(fig_scatter), use_container_width=True)

                if return_figs:
                    return figs_out
                return

    # ── 滞在時間データがない場合はダミーで参考イメージ ──
    if not return_figs:
        st.markdown("**📊 滞在時間別 客単価・時間客単価（参考イメージ）**")
        st.warning(
            "⚠️ **ダミーデータによる分析イメージ**  \n"
            "来店時間・退店時間のデータが不足しているため、ダミーで可視化しています。"
        )

    stay_bins  = ["0-30分", "30-60分", "60-90分", "90-120分", "120-150分", "150分超"]
    avg_spend  = [2200, 3400, 4800, 5800, 6500, 7200]
    hourly_usp = [5280, 3400, 2880, 2640, 2340, 2160]
    cnt_dummy  = [80, 180, 260, 200, 130, 60]

    fig_dummy, (ax1, ax2) = plt.subplots(1, 2, figsize=(11, 4))
    ax1.bar(stay_bins, avg_spend, color="#5b9bd5", edgecolor="white")
    ax1.set_xlabel("滞在時間帯")
    ax1.set_ylabel("平均客単価（円）")
    ax1.set_title("滞在時間帯別 平均客単価 ※ダミーデータ（居酒屋想定）")
    ax1.tick_params(axis="x", rotation=25)

    ax2.plot(stay_bins, hourly_usp, marker="o", color="#e74c3c", linewidth=2)
    ax2.fill_between(range(len(stay_bins)), hourly_usp, alpha=0.2, color="#e74c3c")
    ax2.set_xticks(range(len(stay_bins)))
    ax2.set_xticklabels(stay_bins, rotation=25, ha="right")
    ax2.set_ylabel("時間客単価（円/時間）")
    ax2.set_title("滞在時間帯別 時間客単価 ※ダミーデータ（居酒屋想定）")

    plt.tight_layout()

    dummy_df = pd.DataFrame({
        "滞在時間帯": stay_bins,
        "平均客単価": avg_spend,
        "時間客単価": hourly_usp,
        "件数": cnt_dummy,
    })

    if return_figs:
        figs_out.append(("滞在時間分析 - 客単価・時間客単価（ダミー）", fig_dummy, dummy_df))
    else:
        st.image(_fig_to_buf(fig_dummy), use_container_width=True)
        _insight_advice_block(
            [
                "滞在時間が長いほど客単価は上昇するが、時間客単価（効率）は短時間帯が最も高い（ダミー例）",
                "0-30分帯の時間客単価: 約5,280円/時 vs 150分超: 約2,160円/時 — 短時間の方が約2.4倍効率的",
                "高客単価と高時間効率は二律背反 — 戦略的に使い分けることが重要",
            ],
            [
                "ピーク時（混雑時間帯）は時間制コース・制限を設けて回転率を優先",
                "90分・120分コースの価格は時間客単価に基づいて見直し、短時間の割安感を出さない設計に",
                "閑散時は長時間滞在を歓迎し、追加注文（デザート・締め料理）で客単価を伸ばす施策を展開",
            ],
        )

    if return_figs:
        return figs_out


# ============================================================
# ===== 初期グラフの取得（Excelエクスポート用） =====
# ============================================================

def _render_initial_figs_for_export(df: pd.DataFrame) -> list[tuple[str, io.BytesIO]]:
    """session_state["graphs"] のコードを再実行して (ラベル, PNG BytesIO) を返す。"""
    graphs = st.session_state.get("graphs", [])
    result = []
    for g in graphs:
        if g.get("source") != "initial":
            continue
        plt.close("all")
        safe_globals = {"pd": pd, "np": np, "plt": plt, "matplotlib": matplotlib, "df": df}
        safe_locals = {}
        try:
            cleaned = sanitize_code(g["code"])
            exec(cleaned, safe_globals, safe_locals)  # noqa: S102
            fig = plt.gcf()
            if _fig_has_visible_content(fig):
                buf = io.BytesIO()
                fig.savefig(buf, format="png", dpi=120, bbox_inches="tight")
                buf.seek(0)
                result.append((g["label"], buf))
        except Exception:
            pass
        finally:
            plt.close("all")
    return result


# ============================================================
# ===== Excelエクスポート =====
# ============================================================

def export_to_excel(df: pd.DataFrame, order_df: pd.DataFrame | None) -> io.BytesIO | None:
    """全グラフとデータをExcelに書き出して BytesIO で返す。"""
    if not OPENPYXL_AVAILABLE:
        return None

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # デフォルトシートを削除

    # ヘッダースタイル
    header_fill = PatternFill(start_color="1a237e", end_color="1a237e", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def _write_header_row(ws, headers: list, row: int = 1):
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx, value=str(header))
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align

    def _write_df_to_sheet(ws, data_df: pd.DataFrame, start_row: int = 1):
        if data_df is None or data_df.empty:
            return
        _write_header_row(ws, list(data_df.columns), row=start_row)
        for r_idx, row_data in enumerate(data_df.itertuples(index=False), start=start_row + 1):
            for c_idx, val in enumerate(row_data, start=1):
                if isinstance(val, float):
                    ws.cell(row=r_idx, column=c_idx, value=round(val, 4))
                else:
                    ws.cell(row=r_idx, column=c_idx, value=val)

    def _paste_fig_to_sheet(ws, fig, anchor: str = "A1", dpi: int = 120):
        """matplotlib Figure を PNG に変換してシートに貼り付ける。"""
        buf = io.BytesIO()
        fig.savefig(buf, format="png", dpi=dpi, bbox_inches="tight")
        buf.seek(0)
        plt.close(fig)
        xl_img = XLImage(buf)
        ws.add_image(xl_img, anchor)

    def _paste_buf_to_sheet(ws, png_buf: io.BytesIO, anchor: str = "A1"):
        """BytesIO PNG をシートに貼り付ける。"""
        png_buf.seek(0)
        xl_img = XLImage(png_buf)
        ws.add_image(xl_img, anchor)

    # ── シート1: 初期分析 ──
    ws_init = wb.create_sheet("初期分析")
    title_cell = ws_init["A1"]
    title_cell.value = "初期LLM自動分析グラフ"
    title_cell.font = Font(bold=True, size=14)
    initial_figs = _render_initial_figs_for_export(df)
    anchors = ["A3", "M3", "A35", "M35"]
    for idx, (label, png_buf) in enumerate(initial_figs[:4]):
        if idx < len(anchors):
            _paste_buf_to_sheet(ws_init, png_buf, anchor=anchors[idx])

    # ── シート2〜7: 各追加分析 ──
    analysis_specs = [
        ("①変数特定", _analysis_1_variable_regression, [order_df]),
        ("②商品特定", _analysis_2_product_regression, [df, order_df]),
        ("③ABC分析", _analysis_3_abc_analysis, [df, order_df]),
        ("④バスケット", _analysis_4_basket_analysis, [df, order_df]),
        ("⑤ヘビーライト", _analysis_5_heavy_light, [df, order_df]),
        ("⑥滞留時間", _analysis_6_stay_time_unit_price, [df, order_df]),
    ]

    for sheet_name, func, args in analysis_specs:
        try:
            results = func(*args, return_figs=True) or []
        except Exception:
            results = []

        ws = wb.create_sheet(sheet_name)
        hdr_cell = ws["A1"]
        hdr_cell.value = sheet_name
        hdr_cell.font = Font(bold=True, size=14)

        img_row = 3
        # 最初のグラフを A3 に配置
        for idx, (title, fig, data_df) in enumerate(results):
            col_letter = "A" if idx % 2 == 0 else "M"
            row_num = img_row + (idx // 2) * 32
            # タイトル行
            ws.cell(row=row_num - 1 if row_num > 2 else row_num, column=1, value=title)
            anchor = f"{col_letter}{row_num}"
            _paste_fig_to_sheet(ws, fig, anchor=anchor)

        # データ表（最初の分析で DataFrame がある場合のみ）
        first_df = next((item[2] for item in results if item[2] is not None), None)
        if first_df is not None:
            last_row = img_row + ((len(results) - 1) // 2 + 1) * 32 + 3 if results else 6
            label_cell = ws.cell(row=last_row, column=1, value="【データ表】")
            label_cell.font = Font(bold=True)
            _write_df_to_sheet(ws, first_df, start_row=last_row + 1)

    # ── BytesIO として返す ──
    buf_out = io.BytesIO()
    wb.save(buf_out)
    buf_out.seek(0)
    return buf_out


# ============================================================
# ===== 統合BIダッシュボード =====
# ============================================================

def show_bi_dashboard(df: pd.DataFrame, order_df: pd.DataFrame | None):
    """全分析を1画面にまとめたプロ仕様のBIダッシュボードを表示する。"""

    # ── CSS注入（ダーク BI テーマ） ──
    st.markdown("""
    <style>
    .bi-kpi-card {
        background: #1a1f2e;
        border: 1px solid #30363d;
        border-radius: 12px;
        padding: 16px;
        text-align: center;
        margin-bottom: 8px;
    }
    .bi-kpi-value {
        font-size: 26px;
        font-weight: bold;
        color: #58a6ff;
        margin: 4px 0;
    }
    .bi-kpi-label {
        font-size: 12px;
        color: #8b949e;
        margin: 0;
    }
    .bi-section-header {
        color: #58a6ff;
        font-size: 15px;
        font-weight: bold;
        border-bottom: 1px solid #30363d;
        padding-bottom: 6px;
        margin-bottom: 10px;
    }
    .bi-insight-box {
        background: #161b22;
        border-left: 3px solid #58a6ff;
        padding: 8px 12px;
        border-radius: 4px;
        font-size: 12px;
        color: #c9d1d9;
        margin-bottom: 10px;
        line-height: 1.6;
    }
    </style>
    """, unsafe_allow_html=True)

    # ── ヘッダー ──
    col_title, col_close = st.columns([6, 1])
    with col_title:
        st.markdown("## 📊 外食データ分析 BI Dashboard")
    with col_close:
        if st.button("✕ 閉じる", key="close_dashboard"):
            st.session_state["show_dashboard"] = False
            st.rerun()

    st.markdown("---")

    # ── KPI算出 ──
    kpi_avg_price = "—"
    kpi_total_orders = "—"
    kpi_a_ratio = "—"
    kpi_heavy_ratio = "—"
    kpi_top_pair = "—"

    if order_df is not None and len(order_df) > 0:
        avg_p = order_df["客単価"].mean()
        kpi_avg_price = f"¥{avg_p:,.0f}"
        kpi_total_orders = f"{len(order_df):,}件"

        # A客割合
        q67 = order_df["客単価"].quantile(0.67)
        a_count = (order_df["客単価"] >= q67).sum()
        kpi_a_ratio = f"{a_count / len(order_df) * 100:.0f}%"

        # ヘビー率
        if "ヘビー数" in order_df.columns and "ライト数" in order_df.columns:
            heavy_orders = (order_df["ヘビー数"] > order_df["ライト数"]).sum()
            kpi_heavy_ratio = f"{heavy_orders / len(order_df) * 100:.0f}%"

        # 共起TOP
        if "商品リスト" in order_df.columns:
            try:
                exclude_kw = ["レジ袋", "袋", "クーポン", "割引", "0円"]
                clean_orders = [
                    [itm for itm in lst if itm and not any(ex in itm for ex in exclude_kw)]
                    for lst in order_df["商品リスト"]
                ]
                clean_orders = [lst for lst in clean_orders if len(lst) >= 2]
                if clean_orders:
                    pair_cnt: Counter = Counter()
                    for lst in clean_orders[:500]:
                        for a, b in combinations(set(lst), 2):
                            pair_cnt[tuple(sorted([a, b]))] += 1
                    if pair_cnt:
                        best_pair = pair_cnt.most_common(1)[0][0]
                        kpi_top_pair = f"{best_pair[0][:6]}×{best_pair[1][:6]}"
            except Exception:
                pass

    # ── KPIカード表示 ──
    kpi_cols = st.columns(5)
    kpi_data = [
        ("平均客単価", kpi_avg_price, "💰"),
        ("総注文数", kpi_total_orders, "🧾"),
        ("A客割合", kpi_a_ratio, "⭐"),
        ("ヘビー率", kpi_heavy_ratio, "🍔"),
        ("共起TOP", kpi_top_pair, "🔗"),
    ]
    for col, (label, value, icon) in zip(kpi_cols, kpi_data):
        with col:
            st.markdown(f"""
            <div class="bi-kpi-card">
                <p class="bi-kpi-label">{icon} {label}</p>
                <p class="bi-kpi-value">{value}</p>
            </div>
            """, unsafe_allow_html=True)

    st.markdown("---")

    # ── 初期LLM分析（KPIカードの直下に配置） ──
    st.markdown('<p class="bi-section-header">🧠 初期LLM自動分析</p>', unsafe_allow_html=True)
    initial_result_text = st.session_state.get("initial_result_text")
    initial_graphs = [g for g in st.session_state.get("graphs", []) if g.get("source") == "initial"]

    if initial_result_text or initial_graphs:
        if initial_result_text:
            with st.expander("📝 LLMによる分析コメント（クリックで展開）", expanded=False):
                st.markdown(initial_result_text)
        if initial_graphs:
            st.markdown("**📈 自動生成グラフ**")
            st.caption("💡 グラフにカーソルを合わせると右上に拡大アイコンが現れます。")
            # ── インライン描画（render_graphs_grid はこの呼び出し箇所より後に定義されるため直接inlineする） ──
            _cols_per_row = 3
            for _gi in range(0, len(initial_graphs), _cols_per_row):
                _batch = initial_graphs[_gi: _gi + _cols_per_row]
                _bi_cols = st.columns(len(_batch))
                for _bi_col, _g in zip(_bi_cols, _batch):
                    with _bi_col:
                        st.markdown(f"**📈 グラフ{_g['id']}：{_g['label']}**")
                        plt.close("all")
                        try:
                            # 基本的なコードサニタイズ（先頭の```python行を除去）
                            _code = _g["code"]
                            _code = re.sub(r"^```(?:python)?\s*\n?", "", _code, flags=re.IGNORECASE)
                            _code = re.sub(r"\n?```\s*$", "", _code)
                            _safe_g = {"pd": pd, "np": np, "plt": plt, "df": df}
                            _safe_l: dict = {}
                            exec(_code, _safe_g, _safe_l)  # noqa: S102
                            _fig = plt.gcf()
                            # コンテンツ存在チェック
                            _has_content = bool(
                                _fig and _fig.axes and any(
                                    ax.lines or ax.patches or ax.collections or ax.images
                                    for ax in _fig.axes
                                )
                            )
                            if _has_content:
                                _buf = io.BytesIO()
                                _fig.savefig(_buf, format="png", dpi=120, bbox_inches="tight")
                                _buf.seek(0)
                                st.image(_buf, use_container_width=True)
                            else:
                                st.warning("グラフが描画されませんでした（データ0件の可能性があります）。")
                        except Exception as _ge:
                            st.error(f"グラフ描画エラー: {_ge}")
                        finally:
                            plt.close("all")
    else:
        st.info("初期分析がありません。メイン画面でデータを取得してください。")

    st.markdown("---")

    # ── 詳細分析グリッド（2列） ──
    st.markdown("### 🔬 詳細分析セクション")
    analysis_items = [
        (
            "① 変数特定（重回帰分析）",
            _analysis_1_variable_regression,
            [order_df],
            "客単価に影響する要因（人数・滞在時間・時間帯など）を重回帰分析で特定します。"
            "回帰係数が大きい変数ほど売上への影響が強く、優先的に改善すべき指標です。"
            "正の係数は売上押し上げ要因、負の係数は売上押し下げ要因を示します。",
        ),
        (
            "② 商品特定（商品別回帰）",
            _analysis_2_product_regression,
            [df, order_df],
            "売上に最も貢献している商品を特定します。"
            "高回転・高単価商品への注力と在庫確保が直接的なROI改善につながります。"
            "低貢献商品はメニュー整理・価格見直しの候補として検討してください。",
        ),
        (
            "③ ABC分析",
            _analysis_3_abc_analysis,
            [df, order_df],
            "売上構成比によりA（上位70%）・B（70〜90%）・C（90〜100%）に商品を分類します。"
            "A商品の品質管理・欠品防止を最優先し、C商品はコスト削減・廃番を検討してください。"
            "一般に全商品の約20%でA商品が形成されます（パレートの法則）。",
        ),
        (
            "④ マーケットバスケット分析",
            _analysis_4_basket_analysis,
            [df, order_df],
            "同時注文されやすい商品の組み合わせ（共起）を分析します。"
            "支持度・信頼度・リフト値が高い組み合わせはセット提案・クロスセルの絶好の機会です。"
            "メニュー配置や「おすすめセット」の設計に直接活用できます。",
        ),
        (
            "⑤ ヘビー/ライト分析",
            _analysis_5_heavy_light,
            [df, order_df],
            "高額注文（ヘビー）と低額注文（ライト）の来店パターンを比較します。"
            "ヘビー客が多い時間帯・曜日・店舗を把握し、来店促進施策（特典・DM等）を集中させることで"
            "収益性の高い客層を効率的に増やせます。",
        ),
        (
            "⑥ 滞在時間×客単価",
            _analysis_6_stay_time_unit_price,
            [df, order_df],
            "滞在時間と消費額の関係性を分析します。"
            "滞在時間が長いほど追加注文が増える傾向があれば、滞留施策（BGM・席快適性向上）が有効です。"
            "逆に短時間高単価の客層を増やす回転率向上施策も選択肢になります。",
        ),
    ]

    for i in range(0, len(analysis_items), 2):
        pair = analysis_items[i:i+2]
        cols = st.columns(len(pair))
        for col, (title, func, args, insight) in zip(cols, pair):
            with col:
                st.markdown(f'<p class="bi-section-header">{title}</p>', unsafe_allow_html=True)
                st.markdown(
                    f'<div class="bi-insight-box">{insight}</div>',
                    unsafe_allow_html=True,
                )
                try:
                    results = func(*args, return_figs=True) or []
                    if results:
                        _title, fig, _ = results[0]
                        buf = _fig_to_buf(fig, dpi=100)
                        st.image(buf, use_container_width=True)
                    else:
                        st.info("データなし")
                except Exception as e:
                    st.warning(f"表示エラー: {e}")


# ============================================================
# ===== 追加分析 メイン関数 =====
# ============================================================

def run_additional_analyses(df: pd.DataFrame):
    """6つの追加分析を実行して表示する。"""

    with st.spinner("注文レベルのデータを構築中…"):
        order_df = _build_order_level_df(df)

    if order_df is None:
        st.error(
            "注文単位のデータを構築できませんでした。"
            "「注文番号」または「伝票番号」と「合計金額(税込)」列が必要です。"
        )
        order_df = None

    # ── 各分析をタブで表示 ──
    tabs = st.tabs([
        "① 変数特定", "② 商品特定", "③ ABC分析",
        "④ バスケット", "⑤ ヘビー/ライト", "⑥ 滞留時間",
    ])

    with tabs[0]:
        st.subheader("① 客単価に影響を与える変数の特定（重回帰分析）")
        _analysis_1_variable_regression(order_df)

    with tabs[1]:
        st.subheader("② 客単価に影響を与えている商品の特定（重回帰分析）")
        _analysis_2_product_regression(df, order_df)

    with tabs[2]:
        st.subheader("③ 顧客グルーピング（ABC分析）+ グループ別商品分析")
        _analysis_3_abc_analysis(df, order_df)

    with tabs[3]:
        st.subheader("④ 一緒に注文されるメニューの特定（マーケットバスケット分析）")
        _analysis_4_basket_analysis(df, order_df)

    with tabs[4]:
        st.subheader("⑤ ヘビー系 vs ライト系 食べ物の客単価傾向")
        _analysis_5_heavy_light(df, order_df)

    with tabs[5]:
        st.subheader("⑥ 滞留時間別客単価（時間客単価算出）")
        _analysis_6_stay_time_unit_price(df, order_df)

    st.markdown("---")

    # ── エクスポート / BI ダッシュボード ボタン ──
    col_exp, col_bi, _ = st.columns([1, 1, 3])

    with col_exp:
        if st.button("📥 Excelエクスポート", key="excel_export_btn"):
            if not OPENPYXL_AVAILABLE:
                st.error("openpyxlがインストールされていません。`pip install openpyxl` を実行してください。")
            else:
                with st.spinner("Excelファイルを生成中…"):
                    try:
                        excel_buf = export_to_excel(df, order_df)
                        if excel_buf:
                            st.download_button(
                                label="💾 ダウンロード（.xlsx）",
                                data=excel_buf,
                                file_name="bi_analysis_report.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key="excel_download_btn",
                            )
                            st.success("Excelファイルの生成が完了しました。")
                    except Exception as e:
                        st.error(f"Excelエクスポート中にエラーが発生しました: {e}")

    with col_bi:
        if st.button("📊 統合BIダッシュボード", key="bi_dashboard_btn", type="primary"):
            st.session_state["show_dashboard"] = True
            st.rerun()


# ============================================================
# ===== 以下: v6.6 から引き継いだコード =====
# ============================================================

def map_weathercode_to_label(code: int) -> str:
    """Open-Meteo の weathercode を『晴れ / 曇り / 雨 / 雪 / その他』に変換する。"""
    if pd.isna(code):
        return "不明"
    code = int(code)
    if code in (0, 1):
        return "晴れ"
    if code in (2, 3, 45, 48):
        return "曇り"
    if code in (51, 53, 55, 56, 57, 61, 63, 65, 66, 67, 80, 81, 82):
        return "雨"
    if code in (71, 73, 75, 77, 85, 86):
        return "雪"
    return "その他"


def load_external_signals_csv(uploaded_file) -> pd.DataFrame | None:
    if uploaded_file is None:
        return None
    try:
        df_sig = pd.read_csv(uploaded_file)
    except Exception as e:
        st.error(f"外部シグナルCSVの読み込みでエラーが発生しました: {e}")
        return None
    if "date" not in df_sig.columns:
        st.error("外部シグナルCSVに 'date' 列がありません。YYYY-MM-DD 形式の日付列 'date' を含めてください。")
        return None
    df_sig["date"] = pd.to_datetime(df_sig["date"], errors="coerce").dt.normalize()
    df_sig = df_sig.dropna(subset=["date"])
    for col in df_sig.columns:
        if col == "date":
            continue
        if df_sig[col].dtype == "object":
            df_sig[col] = pd.to_numeric(df_sig[col], errors="coerce")
    return df_sig


SHOW_HISTORY_DASHBOARD = False

# 日本語フォント設定
# Streamlit Cloud (Linux): packages.txt で fonts-noto-cjk をインストール → "Noto Sans CJK JP"
# ローカル Windows: Meiryo / Yu Gothic にフォールバック
import matplotlib.font_manager as _fm

# fontManager.__init__() を直接呼ぶと Python 3.14+ で内部キャッシュが壊れてフォント解決エラーになる。
# 代わりに新しい FontManager インスタンスを生成して差し替える（正しい再構築方法）。
try:
    _fm.fontManager = _fm.FontManager()
except Exception:
    pass  # 失敗しても既存インスタンスをそのまま使う

_FONT_CANDIDATES = [
    "Noto Sans CJK JP", "Noto Sans JP", "NotoSansCJK-Regular",
    "IPAexGothic", "IPAGothic",
    "Meiryo", "Yu Gothic", "MS Gothic",
    "DejaVu Sans",  # matplotlib 同梱の確実なフォント（英字のみ）
]
_available_fonts = {f.name for f in _fm.fontManager.ttflist}
_chosen_font = next((_f for _f in _FONT_CANDIDATES if _f in _available_fonts), "sans-serif")
matplotlib.rcParams["font.family"] = _chosen_font
matplotlib.rcParams["axes.unicode_minus"] = False

BASE_DIR = Path(__file__).resolve().parent
env_path = BASE_DIR / ".env"
load_dotenv(dotenv_path=env_path, override=True)

api_key = os.getenv("OPENAI_API_KEY")
if not api_key:
    raise RuntimeError("OPENAI_API_KEY が .env に設定されていません。")

client = OpenAI(api_key=api_key)
MODEL_NAME = os.getenv("OPENAI_MODEL", "gpt-4.1-mini-2025-04-14")

st.set_page_config(page_title="LLM BI アシスタント", layout="wide")

# Streamlit の管理UIを非表示（Manage App・ハンバーガーメニュー・フッター）
st.markdown("""
<style>
[data-testid="stToolbar"]       { display: none !important; }
[data-testid="manage-app-button"] { display: none !important; }
#MainMenu                       { display: none !important; }
footer                          { display: none !important; }
header                          { display: none !important; }
</style>
""", unsafe_allow_html=True)

st.title("LLM BI アシスタント")
st.caption("v7.0")

# ── セッション状態初期化 ──
for key, default in [
    ("df", None), ("summary_text", None), ("chat_history", []),
    ("initial_result_text", None), ("graphs", []),
    ("next_graph_id", 1), ("uploaded_filename", None),
    ("show_additional", False), ("show_dashboard", False),
    ("dataset", "cafe"),
]:
    if key not in st.session_state:
        st.session_state[key] = default

# ── ダッシュボード表示分岐 ──
if st.session_state.get("show_dashboard"):
    _df_dash = st.session_state.get("df")
    _order_df_dash = _build_order_level_df(_df_dash) if _df_dash is not None else None
    show_bi_dashboard(_df_dash if _df_dash is not None else pd.DataFrame(), _order_df_dash)
    st.stop()


# ── ヘルパー関数 ──
def build_data_summary(df: pd.DataFrame, filename: str, max_rows: int = 5) -> str:
    buf = []
    buf.append(f"ファイル名: {filename}")
    buf.append(f"行数: {len(df):,}, 列数: {len(df.columns)}")
    # 日付範囲
    for dt_col in ["来店時間", "注文日時"]:
        if dt_col in df.columns:
            try:
                dt = pd.to_datetime(df[dt_col], errors="coerce", utc=True).dt.tz_convert("Asia/Tokyo")
                dt_min = dt.dropna().min()
                dt_max = dt.dropna().max()
                if pd.notna(dt_min) and pd.notna(dt_max):
                    buf.append(f"期間: {dt_min.date()} 〜 {dt_max.date()}")
            except Exception:
                try:
                    dt = pd.to_datetime(df[dt_col], errors="coerce")
                    dt_min = dt.dropna().min()
                    dt_max = dt.dropna().max()
                    if pd.notna(dt_min) and pd.notna(dt_max):
                        buf.append(f"期間: {dt_min.date()} 〜 {dt_max.date()}")
                except Exception:
                    pass
            break
    # 店舗一覧
    if "店舗名" in df.columns:
        stores = df["店舗名"].dropna().unique().tolist()
        buf.append(f"店舗({len(stores)}店): {', '.join(map(str, stores[:20]))}")
    # 商品一覧（全種）
    if "商品名" in df.columns:
        items = df["商品名"].dropna().unique().tolist()
        buf.append(f"商品種類数: {len(items)}")
        buf.append(f"商品一覧（全{len(items)}種）: {', '.join(map(str, items))}")
    # 列情報 + 数値列の統計
    buf.append("\n【列情報】")
    for col in df.columns:
        non_null = df[col].notna().sum()
        line = f"- {col}: dtype={df[col].dtype}, 非null={non_null}"
        if df[col].dtype in [np.float64, np.int64, float, int]:
            try:
                s = pd.to_numeric(df[col], errors="coerce").dropna()
                if len(s) > 0:
                    line += f", mean={s.mean():,.1f}, min={s.min():,.0f}, max={s.max():,.0f}"
            except Exception:
                pass
        buf.append(line)
    # サンプル
    buf.append(f"\n【先頭 {max_rows} 行のサンプル（CSV形式）】")
    buf.append(df.head(max_rows).to_csv(index=False))
    return "\n".join(buf)


def _get_unique_values(df: pd.DataFrame, col: str) -> list[str]:
    if col not in df.columns:
        return []
    return df[col].dropna().astype(str).unique().tolist()


def normalize_text_for_matching(text: str) -> str:
    import unicodedata
    if text is None:
        return ""
    s = str(text)
    out = []
    for ch in s:
        if "\u3041" <= ch <= "\u3096":
            out.append(chr(ord(ch) + 0x60))
        else:
            out.append(ch)
    s = "".join(out)
    s = unicodedata.normalize("NFKC", s).lower()
    s = s.replace("〜", "ー").replace("~", "ー")
    s = re.sub(r"[\s　]", "", s)
    return s


def enhanced_fuzzy_match(query: str, candidates: list[str]) -> str | None:
    if not query or not candidates:
        return None
    qn = normalize_text_for_matching(query)
    for c in candidates:
        if qn == normalize_text_for_matching(c):
            return c
    best, best_score = None, 0.0
    for c in candidates:
        cn = normalize_text_for_matching(c)
        if qn in cn or cn in qn:
            score = min(len(qn), len(cn)) / max(len(qn), len(cn))
            if score > best_score:
                best_score = score
                best = c
    if best is not None and best_score >= 0.45:
        return best
    import difflib
    norm_candidates = [normalize_text_for_matching(c) for c in candidates]
    hit = difflib.get_close_matches(qn, norm_candidates, n=1, cutoff=0.35)
    if hit:
        idx = norm_candidates.index(hit[0])
        return candidates[idx]
    return None


_INFO_QUERY_PATTERNS = re.compile(
    r"は何がありますか|何がある|を教えて|一覧|リスト|どんな.{0,6}(商品|メニュー|店舗)|"
    r"(商品|メニュー|店舗).{0,6}(何|どんな|どのような|種類)|どのような|何種類|何店舗"
)

def build_fuzzy_context_for_chat(df: pd.DataFrame, user_text: str) -> tuple[str, list[str], dict]:
    # 一覧確認・情報質問はfuzzyマッチをスキップ（誤補正防止）
    if _INFO_QUERY_PATTERNS.search(user_text):
        return user_text, [], {"extra_system": "", "stores": {}, "products": {}}

    notes: list[str] = []
    resolved = {"stores": {}, "products": {}}
    store_candidates = _get_unique_values(df, "店舗名")
    product_candidates = _get_unique_values(df, "商品名")
    store_tokens = []
    store_tokens.extend(re.findall(r"([^\s「」『』,。、]+店)", user_text))
    store_tokens.extend(re.findall(r"([^\s「」『』,。、]+)(?:の売上|について|の分析|を知りたい|の売り上げ)", user_text))
    location_patterns = [
        r"([あ-んア-ン一-龥]+ぷら[ー〜]*ざ[ー〜]*)",
        r"([あ-んア-ン一-龥]+プラ[ー〜]*ザ[ー〜]*)",
        r"([あ-んア-ン一-龥]+がおか)",
        r"([あ-んア-ン一-龥]+ばし)",
        r"([あ-んア-ン一-龥]+だ)",
    ]
    for pattern in location_patterns:
        store_tokens.extend(re.findall(pattern, user_text))
    product_tokens: list[str] = []
    for pat in [
        # 居酒屋向けメニューパターン
        r"([^\s「」『』]+焼き鳥)",
        r"([^\s「」『』]+唐揚げ)",
        r"([^\s「」『』]+枝豆)",
        r"([^\s「」『』]+刺身)",
        r"([^\s「」『』]+ビール)",
        r"([^\s「」『』]+ハイボール)",
        r"([^\s「」『』]+サワー)",
        r"([^\s「」『』]+チューハイ)",
        r"([^\s「」『』]+串焼き)",
        r"([^\s「」『』]+鍋)",
        r"([^\s「」『』]+ドリンク)",
        r"([^\s「」『』]+セット)",
    ]:
        product_tokens.extend(re.findall(pat, user_text))
    for tok in store_tokens:
        tok = tok.strip()
        if not tok:
            continue
        best = enhanced_fuzzy_match(tok, store_candidates)
        if best and best != tok:
            resolved["stores"][tok] = best
    for tok in product_tokens:
        tok = tok.strip()
        if not tok:
            continue
        best = enhanced_fuzzy_match(tok, product_candidates)
        if best and best != tok:
            resolved["products"][tok] = best
    auto_lines: list[str] = []
    if resolved["stores"]:
        notes.append("曖昧な店舗名を次のように補正しました：")
        auto_lines.append("【店舗名の補正】次の正式名称として扱って集計してください。")
        for src, dst in resolved["stores"].items():
            notes.append(f"  『{src}』→『{dst}』")
            auto_lines.append(f"- '{src}' は '{dst}' として扱う")
    if resolved["products"]:
        notes.append("曖昧な商品名を次のように補正しました：")
        auto_lines.append("【商品名の補正】次の正式名称として扱って集計してください。")
        for src, dst in resolved["products"].items():
            notes.append(f"  『{src}』→『{dst}』")
            auto_lines.append(f"- '{src}' は '{dst}' として扱う")
    patched_user_text = user_text
    if auto_lines:
        patched_user_text = user_text + "\n\n[AUTO_ANNOTATION]\n" + "\n".join(auto_lines)
    extra_system = ""
    if auto_lines:
        extra_system = "【AUTO_ANNOTATIONがある場合は必ずそれを優先】\n" + "\n".join(auto_lines)
    return patched_user_text, notes, {"extra_system": extra_system, **resolved}


def parse_llm_response(text: str) -> tuple[str, list[str]]:
    pattern = r"```python(.*?)```"
    codes = [m.strip() for m in re.findall(pattern, text, re.DOTALL)]
    text_wo_codes = re.sub(r"```python.*?```", "", text, flags=re.DOTALL).strip()
    lines = [
        ln for ln in text_wo_codes.splitlines()
        if "グラフ描画用Pythonコード" not in ln
        and "matplotlibコード" not in ln
        and "matplotlib 用Pythonコード" not in ln
    ]
    return "\n".join(lines).strip(), codes


def sanitize_code(code: str) -> str:
    import re as _re
    cleaned_lines = []
    for line in code.splitlines():
        stripped = line.strip()
        if stripped.startswith("import ") or stripped.startswith("from "):
            continue
        # LLMがフォントを上書きする行を除去（文字化け防止）
        if "rcParams" in stripped and "font" in stripped:
            continue
        # LLMがローカルCSV/Excelを読もうとする行を除去（df は既にメモリ上にある）
        if _re.search(r'pd\.read_csv\s*\(|pd\.read_excel\s*\(|open\s*\(', stripped):
            cleaned_lines.append("# [removed: file read blocked — use df variable]")
            continue
        cleaned_lines.append(line)
    code = "\n".join(cleaned_lines)
    # pandas 2.0+ で廃止された .append() を pd.concat() に自動変換
    code = _re.sub(
        r'(\w+)\.append\((\w+),\s*ignore_index\s*=\s*True\)',
        r'pd.concat([\1, \2], ignore_index=True)',
        code,
    )
    # tick_params() の ha= は無効パラメータ → 除去（matplotlib 3.x の変更）
    # 正しくは plt.xticks(ha=...) または set_xticklabels(ha=...) を使う
    code = _re.sub(
        r'(tick_params\([^)]*?),\s*ha\s*=\s*[\'"][^\'"]*[\'"]([^)]*?\))',
        r'\1\2',
        code,
    )
    # 先頭に日本語フォント強制設定を挿入（起動時に決定した _chosen_font を使う）
    font_fix = (
        f"plt.rcParams['font.family'] = {_chosen_font!r}\n"
        "plt.rcParams['axes.unicode_minus'] = False\n"
    )
    return font_fix + code


def _fig_has_visible_content(fig) -> bool:
    try:
        if fig is None:
            return False
        axes = getattr(fig, "axes", None) or []
        if not axes:
            return False
        for ax in axes:
            if getattr(ax, "lines", None) and len(ax.lines) > 0:
                return True
            if getattr(ax, "patches", None) and len(ax.patches) > 0:
                return True
            if getattr(ax, "collections", None) and len(ax.collections) > 0:
                return True
            if getattr(ax, "images", None) and len(ax.images) > 0:
                return True
        return False
    except Exception:
        return True


def render_graph(graph: dict, df: pd.DataFrame):
    gid = graph["id"]
    label = graph["label"]
    code = graph["code"]
    st.markdown(f"**📈 グラフ{gid}：{label}**")
    plt.close("all")
    safe_globals = {"pd": pd, "np": np, "plt": plt, "matplotlib": matplotlib, "df": df}
    safe_locals = {}
    try:
        cleaned = sanitize_code(code)
        exec(cleaned, safe_globals, safe_locals)  # noqa: S102
        fig = plt.gcf()
        if _fig_has_visible_content(fig):
            buf = io.BytesIO()
            fig.savefig(buf, format="png", dpi=150, bbox_inches="tight")
            buf.seek(0)
            st.image(buf, use_container_width=True)
        else:
            st.warning("グラフが描画されませんでした（データ0件の可能性があります）。店舗名・期間・条件を変えて試してください。")
            with st.expander("🔎 ヒント"):
                st.write("条件で絞り込み後の件数が0件だと、グラフが空になります。店舗名の表記揺れや期間を確認してください。")
    except Exception as e:
        err_str = str(e)
        tb_str = "".join(traceback.format_exception(type(e), e, e.__traceback__))
        # 空 DataFrame を plot したときの IndexError を専用メッセージで案内
        if isinstance(e, IndexError) and "size 0" in err_str:
            st.warning(
                "⚠️ **グラフを描画するデータが0件でした。**\n\n"
                "考えられる原因:\n"
                "- 指定した店舗・商品・期間に一致するデータが存在しない\n"
                "- 前年比分析で比較する年の一方にデータがない\n\n"
                "**対処方法:** チャットで「{店舗名} のデータが存在する期間を教えて」と聞いてから再度お試しください。"
            )
        else:
            st.error("グラフ描画中にエラーが発生しました。コードや指示を調整して再度お試しください。")
        with st.expander(f"🔧 エラー詳細（グラフ{gid} / 開発者向け）"):
            st.text(tb_str)
    finally:
        plt.close("all")


def render_graphs_grid(graphs: list, df: pd.DataFrame, cols_per_row: int = 3):
    for i in range(0, len(graphs), cols_per_row):
        batch = graphs[i : i + cols_per_row]
        cols = st.columns(len(batch))
        for col, graph in zip(cols, batch):
            with col:
                render_graph(graph, df)



_GRAPH_QUALITY_RULES = (
    "【グラフ品質ルール（必須）】\n"
    "・全ての金額軸のラベルには「（円）」を付ける。例: ax.set_ylabel(\"平均客単価（円）\")\n"
    "・金額の数値表示は必ず {:,} 形式でカンマ区切りにする。例: f\"{val:,.0f}円\"\n"
    "・全グラフに必ず: タイトル(set_title)・X軸ラベル(set_xlabel)・Y軸ラベル(set_ylabel) を設定する\n"
    "・バーグラフには各バーに数値アノテーションを付ける\n"
    "・天気列(temperature_2m_max/min/mean, precipitation_sum, weather_label)が存在する場合は必ずそれを使ったグラフを1つ作成する\n"
    "\n【空データ対策（必須）】\n"
    "・絞り込み後は必ず len() で件数を確認する。0件の場合は .plot() を呼ばず、"
    "ax.text(0.5, 0.5, 'データなし', ha='center', va='center', fontsize=14) で代替表示すること\n"
    "・前年比・年別比較では pivot/unstack 後に全カラムが揃っているか確認し、"
    "不足年がある場合は fillna(0) するか存在する年のみでグラフを描く\n"
    "・DataFrame.plot() / Series.plot() を呼ぶ直前に if data.empty: でガードすること\n"
    "\n【データ列の意味】\n"
    "・合計金額(税込): 伝票合計（客単価）\n"
    "・単価: 1商品の販売価格\n"
    "・数量: 商品数量\n"
    "・人数: 1来店グループの人数\n"
    "・来店時間・退店時間: 実際の入退店時刻（差分が滞在時間）\n"
    "・客層: 顧客属性（例: 一般客、VIPなど）\n"
    "・商品名: 注文された商品名\n"
    "・通貨単位: 日本円（JPY）\n"
)

def call_llm_initial(summary_text: str, user_prompt: str) -> str:
    system_prompt = (
        "あなたは『売上データ分析専用』のアシスタントです。\n"
        "アップロードされたCSVデータのサマリーは以下の通りです。\n"
        "==== データサマリー ====\n"
        f"{summary_text}\n"
        "=======================\n\n"
        "ルール:\n"
        "・必ず『複数（2〜4個）のグラフ』を作り、それぞれに簡潔な解釈（気づき）を添える。\n"
        "・ユーザーに再確認を求めず、dfから自分で集計して描画する。\n"
        "・店舗別などカテゴリが多すぎる場合は、売上上位10〜20に絞る。\n"
        "・もし次の列が存在する場合は、最低1つはそれを使ったグラフを含める：\n"
        "  - temperature_2m_mean / temperature_2m_max / temperature_2m_min\n"
        "  - weathercode / weather_label\n"
        "  - precipitation_sum など降水系\n"
        "  - event_score / sns_score など外部シグナル\n"
        "・グラフ描画コードは必ず ```python ... ``` に入れる（matplotlib）。\n"
        "【コードのルール（重要）】\n"
        "・pandas 2.0+ を使用中のため Series.append() / DataFrame.append() は廃止。必ず pd.concat() を使うこと。\n"
        "・np（numpy）は利用可能。\n"
        + _GRAPH_QUALITY_RULES
    )
    resp = client.chat.completions.create(
        model=MODEL_NAME,
        messages=[
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt},
        ],
        temperature=0.2,
    )
    return resp.choices[0].message.content


def call_llm_chat(summary_text: str, chat_history: list[dict], extra_system: str = "") -> str:
    system_prompt = (
        "あなたは売上データに関するチャットアシスタントです。\n"
        "取得済みの売上データのサマリーは以下の通りです。\n"
        "==== データサマリー ====\n"
        f"{summary_text}\n"
        "=======================\n\n"
        "【絶対禁止事項】\n"
        "・pd.read_csv() / pd.read_excel() / open() などでファイルを読み込んではいけない\n"
        "・データは既に変数 df としてメモリ上に読み込まれている。必ず df をそのまま使うこと\n\n"
        "【質問の種類による回答方針】\n"
        "■ グラフ不要な質問（以下に該当する場合）はテキストのみで回答し、コードブロックは出力しないこと：\n"
        "  - 商品一覧・店舗一覧の確認（「〜は何がありますか」「〜を教えて」など）\n"
        "  - データの件数・構造・定義の確認\n"
        "  - 「セット商品は？」「どんなメニューがある？」などの一覧確認\n"
        "  - この場合、[AUTO_ANNOTATION] の補正情報は無視してよい\n"
        "■ グラフが有用な質問（集計・比較・トレンド分析など）は以下のルールでグラフを作成：\n"
        "  - 1〜3個のグラフを作成し、気づき（箇条書き2〜3個）と追加分析案（1〜2個）を含める\n"
        "  - 前年比・年別比較など複数年を扱う場合は年ごとに分けた複数グラフを作ること\n"
        "  - 店舗別などカテゴリが多すぎる場合は売上上位10〜20に絞る\n"
        "  - グラフ描画コードは必ず ```python ... ``` に入れる\n"
        "・ユーザーに再確認を求めてはいけない。dfから自分で件数や集計を計算する。\n"
        "・メッセージ末尾に [AUTO_ANNOTATION] が付くことがある（グラフ分析時のみ最優先で使用）。\n"
        "【店舗/商品フィルタのルール（グラフ分析時）】\n"
        "・[AUTO_ANNOTATION] に正式名称がある場合は、まず df の該当列で『==（完全一致）』で絞る。\n"
        "・[AUTO_ANNOTATION] が無い場合のみ contains を許可。\n"
        "・存在する場合は weather_label / weathercode / temperature_2m_* / precipitation_sum / event_score / sns_score を優先的に使った分析も提案する。\n"
        "【コードのルール（重要）】\n"
        "・pandas 2.0+ を使用中のため Series.append() / DataFrame.append() は廃止。必ず pd.concat() を使うこと。\n"
        "・np（numpy）は利用可能。\n"
        + _GRAPH_QUALITY_RULES
    )
    if extra_system:
        system_prompt += "\n\n" + extra_system
    messages = [{"role": "system", "content": system_prompt}]
    for m in chat_history:
        if isinstance(m, dict) and "role" in m and "content" in m:
            messages.append({"role": m["role"], "content": m["content"]})
    resp = client.chat.completions.create(
        model=MODEL_NAME,
        messages=messages,
        temperature=0.2,
    )
    return resp.choices[0].message.content


def fetch_weather_daily(lat: float, lon: float, start: date, end: date) -> pd.DataFrame:
    url = "https://archive-api.open-meteo.com/v1/era5"
    params = {
        "latitude": lat, "longitude": lon,
        "start_date": start.strftime("%Y-%m-%d"),
        "end_date": end.strftime("%Y-%m-%d"),
        "daily": (
            "temperature_2m_max,temperature_2m_min,temperature_2m_mean,"
            "precipitation_sum,weathercode"
        ),
        "timezone": "Asia/Tokyo",
    }
    resp = requests.get(url, params=params, timeout=10)
    resp.raise_for_status()
    data = resp.json()
    daily = data.get("daily", {})
    dates = daily.get("time", [])
    if not dates:
        return pd.DataFrame()
    df_w = pd.DataFrame({
        "date": pd.to_datetime(dates),
        "temperature_2m_max": daily.get("temperature_2m_max", []),
        "temperature_2m_min": daily.get("temperature_2m_min", []),
        "temperature_2m_mean": daily.get("temperature_2m_mean", []),
        "precipitation_sum": daily.get("precipitation_sum", []),
        "weathercode": daily.get("weathercode", []),
    })
    df_w["date"] = df_w["date"].dt.normalize()
    df_w["weather_label"] = df_w["weathercode"].apply(map_weathercode_to_label)
    return df_w


# ============================================================
# ===== サマリーキャッシュ ヘルパー =====
# ============================================================

_TIMEBAND_ORDER = ["〜17時(昼)", "17〜20時(夕方)", "20〜23時(夜)", "23時〜(深夜)"]


def _load_summary_cache_from_db() -> dict:
    """Supabase の summary_cache テーブルからキャッシュを読み込む。失敗時は空dictを返す。"""
    if not SUPABASE_AVAILABLE:
        return {}
    try:
        _sb = get_supabase_client()
        res = _sb.table("summary_cache").select("*").eq("id", 1).execute()
        if res.data:
            row = res.data[0]
            return {
                "generated_at":  row.get("generated_at", "不明"),
                "total_visits":  row.get("total_visits", 0),
                "store_month":   row.get("store_month", []),
                "store_timeband": row.get("store_timeband", []),
            }
    except Exception:
        pass
    return {}


def _rebuild_summary_cache() -> dict:
    """visits テーブルからキャッシュを再集計し Supabase の summary_cache テーブルに保存する。"""
    from datetime import datetime as _dt

    if not SUPABASE_AVAILABLE:
        st.error("supabase_loader が利用できません。")
        return {}

    try:
        _sb = get_supabase_client()
        df = fetch_visits_for_summary(_sb)
    except Exception as e:
        st.error(f"DB 取得エラー: {e}")
        return {}

    if df.empty:
        st.warning("visits テーブルにデータがありません。")
        return {}

    def _to_band(h):
        if h < 17: return "〜17時(昼)"
        if h < 20: return "17〜20時(夕方)"
        if h < 23: return "20〜23時(夜)"
        return "23時〜(深夜)"

    dt_ser = pd.to_datetime(df["visit_time"], format="ISO8601", errors="coerce", utc=True)
    jst = dt_ser.dt.tz_convert("Asia/Tokyo")
    df = df.copy()
    df["month"]     = jst.dt.strftime("%Y-%m")
    df["hour"]      = jst.dt.hour
    df["time_band"] = df["hour"].apply(_to_band)
    df = df.dropna(subset=["month"])

    # 未来月・異常日付を除外
    _today_month = pd.Timestamp.now(tz="Asia/Tokyo").strftime("%Y-%m")
    _before = len(df)
    df = df[df["month"] <= _today_month]
    _dropped = _before - len(df)
    if _dropped > 0:
        st.warning(f"⚠️ 未来日付のレコードを {_dropped} 件除外しました（visits テーブルに異常データあり）。")

    store_month = (
        df.groupby(["store_name", "month"], sort=True)["receipt_no"]
        .nunique().reset_index()
        .rename(columns={"store_name": "店舗名", "receipt_no": "伝票数"})
    )
    store_timeband = (
        df.groupby(["store_name", "time_band"], sort=True)["receipt_no"]
        .nunique().reset_index()
        .rename(columns={"store_name": "店舗名", "receipt_no": "伝票数"})
    )

    cache = {
        "generated_at":  _dt.now().strftime("%Y-%m-%d %H:%M:%S"),
        "total_visits":  int(len(df)),
        "store_month":   store_month.to_dict(orient="records"),
        "store_timeband": store_timeband.to_dict(orient="records"),
    }

    # Supabase に upsert（id=1 の行を常に上書き）
    try:
        _sb.table("summary_cache").upsert({
            "id":            1,
            "generated_at":  cache["generated_at"],
            "total_visits":  cache["total_visits"],
            "store_month":   cache["store_month"],
            "store_timeband": cache["store_timeband"],
        }).execute()
    except Exception as e:
        st.warning(f"キャッシュの保存に失敗しました（読み取りは可能です）: {e}")

    return cache


def _show_summary_cache():
    """Supabase の summary_cache テーブルを読んでヒートマップ表示する。"""
    cache: dict = _load_summary_cache_from_db()

    col_info, col_btn = st.columns([4, 1])
    with col_info:
        if cache and cache.get("generated_at", "未生成") != "未生成":
            total_v = cache.get("total_visits", "")
            total_str = f"  |  来店記録: {total_v:,}件" if isinstance(total_v, int) else ""
            st.caption(
                f"生成日時: {cache.get('generated_at', '不明')}  |  "
                f"データソース: DB（visits テーブル）{total_str}"
            )
        else:
            st.caption("キャッシュが未生成です。右の「🔄 再生成」ボタンを押してください（DBへ接続します）。")

    with col_btn:
        if st.button("🔄 キャッシュを再生成", key="rebuild_cache_btn"):
            with st.spinner("集計中..."):
                cache = _rebuild_summary_cache()
            if cache:
                st.success("✅ 再生成しました。")
            else:
                st.error("生成に失敗しました。")

    if not cache:
        return

    # ── 店舗×月 ヒートマップ ────────────────────────────────────────
    st.markdown("#### 🗓️ 店舗 × 月　伝票数")
    df_sm = pd.DataFrame(cache.get("store_month", []))
    if not df_sm.empty:
        pivot_sm = (
            df_sm.pivot(index="店舗名", columns="month", values="伝票数")
            .fillna(0).astype(int)
        )
        # 列を昇順ソート
        pivot_sm = pivot_sm.reindex(sorted(pivot_sm.columns), axis=1)
        styled_sm = (
            pivot_sm.style
            .background_gradient(cmap="Greens", axis=None)
            .format("{:,}")
        )
        st.dataframe(styled_sm, use_container_width=True)
    else:
        st.info("店舗×月データがありません。")

    # ── 店舗×時間帯 ヒートマップ ─────────────────────────────────────
    st.markdown("#### ⏰ 店舗 × 時間帯　伝票数（全月合計）")
    df_st = pd.DataFrame(cache.get("store_timeband", []))
    if not df_st.empty:
        # 時間帯を定義順にソート
        present_bands = [b for b in _TIMEBAND_ORDER if b in df_st["time_band"].values]
        df_st["time_band"] = pd.Categorical(
            df_st["time_band"], categories=present_bands, ordered=True
        )
        pivot_st = (
            df_st.pivot(index="店舗名", columns="time_band", values="伝票数")
            .fillna(0).astype(int)
        )
        styled_st = (
            pivot_st.style
            .background_gradient(cmap="YlOrRd", axis=None)
            .format("{:,}")
        )
        st.dataframe(styled_st, use_container_width=True)
    else:
        st.info("店舗×時間帯データがありません。")


# ============================================================
# ===== UI =====
# ============================================================

# ── データ概要サマリー（非表示）──────────────────────────────────────────────
# with st.expander("📊 データ概要サマリー（事前集計）", expanded=False):
#     _show_summary_cache()

st.divider()

# ============================================================
# ===== DB データ取得 UI =====
# ============================================================

st.markdown("### 📅 DB からデータを取得")

# ── データセット選択 ──────────────────────────────────────────────
DATASET_CONFIG = {
    "izakaya": {
        "rpc_name":     "get_izakaya_sales",
        "stores_table": "stores",
        "visits_table": "visits",
        "label":        "居酒屋",
    },
    "cafe": {
        "rpc_name":     "get_cafe_sales",
        "stores_table": "cafe_stores",
        "visits_table": "cafe_visits",
        "label":        "カフェ（Cafe Bloom）",
    },
}

def _on_dataset_change():
    for _k in ["df", "summary_text", "chat_history", "initial_result_text",
               "graphs", "next_graph_id", "uploaded_filename", "show_additional",
               "sb_stores_df_izakaya", "sb_stores_df_cafe"]:
        st.session_state.pop(_k, None)

_ds_idx = list(DATASET_CONFIG.keys()).index(st.session_state.get("dataset", "cafe"))
dataset_choice = st.selectbox(
    "📊 データセット",
    options=list(DATASET_CONFIG.keys()),
    format_func=lambda k: ("🍺 " if k == "izakaya" else "☕ ") + DATASET_CONFIG[k]["label"],
    index=_ds_idx,
    on_change=_on_dataset_change,
)
st.session_state["dataset"] = dataset_choice
ACTIVE_DATASET = DATASET_CONFIG[dataset_choice]

if not SUPABASE_AVAILABLE:
    st.error("supabase_loader モジュールが見つかりません。`pip install supabase` を実行してください。")
else:
    # 取得可能月リスト（データセットごとにキャッシュ）
    FALLBACK_MONTHS = ["2024-09", "2024-10", "2025-09", "2025-10"]

    @st.cache_data(ttl=300)
    def _get_available_months(visits_table: str):
        try:
            _sb = get_supabase_client()
            months = fetch_available_months(_sb, visits_table=visits_table)
            return months if months else FALLBACK_MONTHS
        except Exception:
            return FALLBACK_MONTHS

    KNOWN_MONTHS = _get_available_months(ACTIVE_DATASET["visits_table"])

    # 店舗リスト取得（データセットごとにキャッシュ）
    _stores_key = f"sb_stores_df_{dataset_choice}"
    if _stores_key not in st.session_state:
        try:
            _sb = get_supabase_client()
            st.session_state[_stores_key] = sb_fetch_stores(_sb, table=ACTIVE_DATASET["stores_table"])
        except Exception as _e:
            st.session_state[_stores_key] = pd.DataFrame()

    df_stores_master = st.session_state.get(_stores_key, pd.DataFrame())

    col_month, col_store = st.columns([1, 1])
    with col_month:
        available = KNOWN_MONTHS
        selected_months = st.multiselect(
            "分析月を選択（複数可）",
            options=available,
            default=available[-2:] if len(available) >= 2 else available,
            key="sb_months",
        )
    with col_store:
        if not df_stores_master.empty and "store_name" in df_stores_master.columns:
            store_options = df_stores_master["store_name"].dropna().tolist()
            selected_store_names = st.multiselect(
                "店舗を絞り込む（空=全店舗）",
                options=store_options,
                key=f"sb_stores_{dataset_choice}",
            )
            if selected_store_names and "store_id" in df_stores_master.columns:
                sb_store_ids = df_stores_master.loc[
                    df_stores_master["store_name"].isin(selected_store_names), "store_id"
                ].tolist()
            else:
                sb_store_ids = None
        else:
            sb_store_ids = None

    if st.button("🔄 DB からデータを取得", type="primary", key="sb_fetch_btn"):
        if not selected_months:
            st.error("分析月を1つ以上選択してください。")
        else:
            _months_label = "・".join(selected_months)
            _sb = get_supabase_client()
            _all_dfs = []
            _error_months = []

            # 全月のチャンク数を事前計算（追加クエリなし・純粋な日付計算）
            _total_chunks = sum(
                count_fetch_chunks(*months_to_date_range([m]))
                for m in selected_months
            )
            _done_chunks = [0]  # リストで可変参照（クロージャ共有用）

            _progress_bar = st.progress(0, text="取得準備中...")
            _status_text  = st.empty()

            for _mi, _month in enumerate(selected_months):
                _s, _e = months_to_date_range([_month])

                def _cb(n, label=_month):
                    _done_chunks[0] += 1
                    _pct = min(_done_chunks[0] / _total_chunks, 1.0)
                    _progress_bar.progress(
                        _pct,
                        text=(
                            f"📡 {label} — "
                            f"{_done_chunks[0]}/{_total_chunks} チャンク完了 "
                            f"({_pct*100:.0f}%) 累計 {n:,} 件"
                        ),
                    )

                try:
                    _df_m = sb_fetch_sales_data(_sb, _s, _e, sb_store_ids, progress_callback=_cb, rpc_name=ACTIVE_DATASET["rpc_name"])
                    if not _df_m.empty:
                        _all_dfs.append(_df_m)
                except Exception as _ex:
                    _error_months.append(f"{_month}: {_ex}")

            _progress_bar.empty()
            _status_text.empty()

            if _error_months:
                st.warning(
                    "一部の月でエラーが発生しました:\n" + "\n".join(_error_months) + "\n\n"
                    f"**解決方法:** `etc/{'cafe_setup.sql' if dataset_choice == 'cafe' else 'supabase_setup.sql'}` を DB SQL Editor で実行してください。\n"
                    f"（インデックス追加 + `{ACTIVE_DATASET['rpc_name']}` RPC 関数の作成 + 権限付与）"
                )

            if _all_dfs:
                _df = pd.concat(_all_dfs, ignore_index=True)
                # 重複除去: 来店時間+伝票番号+メニューの組み合わせで判定
                # ※伝票番号は店舗・日付をまたいで同じ番号が使われることがあるため
                #   来店時間を含めることで異なる来店の同番号を別伝票として正しく扱う
                _dedup_cols = [c for c in ["来店時間", "伝票番号", "商品名"] if c in _df.columns]
                if _dedup_cols:
                    _df = _df.drop_duplicates(subset=_dedup_cols)

                st.session_state["df"] = _df
                st.session_state["uploaded_filename"] = f"supabase_{_months_label}.csv"
                st.session_state["summary_text"] = None
                st.session_state["chat_history"] = []
                st.session_state["initial_result_text"] = None
                st.session_state["graphs"] = []
                st.session_state["next_graph_id"] = 1
                st.session_state["show_additional"] = True

                # 伝票数: 来店時間+伝票番号の組み合わせをユニークカウント
                # （伝票番号単独では複数月で重複する場合があるため）
                if "来店時間" in _df.columns and "伝票番号" in _df.columns:
                    _visits_cnt = _df.drop_duplicates(subset=["来店時間", "伝票番号"]).shape[0]
                elif "伝票番号" in _df.columns:
                    _visits_cnt = _df["伝票番号"].nunique()
                else:
                    _visits_cnt = "-"
                st.success(
                    f"✅ {len(_df):,} 件の明細データを取得しました"
                    f"（伝票数: {_visits_cnt:,}件、期間: {_months_label}）"
                )
                with st.expander("📋 取得データのプレビュー（先頭5行）"):
                    st.dataframe(_df.head())
            elif not _error_months:
                st.error(
                    "データが取得できませんでした。\n\n"
                    "**考えられる原因:**\n"
                    "- DB の RLS ポリシーで読み取りが制限されている\n"
                    "- 指定期間にデータが存在しない\n\n"
                    "**RLS 解決方法:** DB SQL Editor で以下を実行してください:\n"
                    "```sql\n"
                    "CREATE POLICY \"anon_read_visits\" ON visits FOR SELECT TO anon USING (true);\n"
                    "CREATE POLICY \"anon_read_orders\" ON orders FOR SELECT TO anon USING (true);\n"
                    "CREATE POLICY \"anon_read_order_items\" ON order_items FOR SELECT TO anon USING (true);\n"
                    "CREATE POLICY \"anon_read_stores\" ON stores FOR SELECT TO anon USING (true);\n"
                    "```"
                )

# ── 現在読み込み済みデータの表示 ──
if st.session_state.get("df") is not None:
    _cur_df = st.session_state["df"]
    _cur_fname = st.session_state.get("uploaded_filename", "")
    st.info(
        f"📊 現在のデータ: **{_cur_fname}** | "
        f"{len(_cur_df):,} 行 | "
        f"伝票数: {_cur_df.drop_duplicates(subset=['来店時間','伝票番号']).shape[0] if ('来店時間' in _cur_df.columns and '伝票番号' in _cur_df.columns) else (_cur_df['伝票番号'].nunique() if '伝票番号' in _cur_df.columns else '-')}件"
    )

# ── 天気データ管理（非表示）──

st.markdown("---")


@st.fragment
def _chat_section():
    st.header("🧠 分析指示チャット（データ分析専用）")

    chat_query = st.text_area(
        "上記と同じデータについて、追加で知りたいこと・試したい集計を自由に入力してください。",
        key="chat_query", height=120,
        placeholder="例）商品別の売り上げ構成を知りたい",
    )

    if st.button("送信する（分析・質問）", key="chat_button"):
        df_chat = st.session_state.get("df")
        if df_chat is None:
            st.error("先に DB からデータを取得してください。")
        elif not chat_query.strip():
            st.error("チャット内容が空です。知りたいことを入力してください。")
        else:
            patched_user_text = chat_query
            fuzzy_notes: list[str] = []
            extra_system = ""
            resolved_info = {}
            try:
                patched_user_text, fuzzy_notes, resolved_info = build_fuzzy_context_for_chat(df_chat, chat_query)
                extra_system = resolved_info.get("extra_system", "")
            except Exception as e:
                st.warning(f"曖昧マッチ処理中にエラーが発生しましたが、チャット自体は続行します: {e}")
            if fuzzy_notes:
                st.info("\n".join(fuzzy_notes))
            try:
                store_map = (resolved_info or {}).get("stores", {})
                if store_map and ("店舗名" in df_chat.columns):
                    src = list(store_map.keys())[0]
                    dst = list(store_map.values())[0]
                    if int((df_chat["店舗名"] == dst).sum()) == 0:
                        cand = _get_unique_values(df_chat, "店舗名")
                        dst_norm = normalize_text_for_matching(dst)
                        hint = [c for c in cand if dst_norm and (dst_norm in normalize_text_for_matching(c))][:10]
                        st.warning("補正後の店舗名でデータが0件でした。候補を選んで再実行してください。")
                        if hint:
                            st.write("候補（クリックでチャット欄に反映）:")
                            for i, cand_name in enumerate(hint, start=1):
                                if st.button(f"候補{i}: {cand_name}", key=f"cand_store_{i}"):
                                    original = st.session_state.get("chat_query", "") or chat_query
                                    new_q = original
                                    if src and (src in new_q):
                                        new_q = new_q.replace(src, cand_name)
                                    elif dst and (dst in new_q):
                                        new_q = new_q.replace(dst, cand_name)
                                    else:
                                        new_q = f"{cand_name} の " + new_q
                                    st.session_state["chat_query"] = new_q
                                    st.rerun()
                        else:
                            st.info("候補を見つけられませんでした。店舗名一覧を確認してください。")
                        st.stop()
            except Exception:
                pass

            summary_text = st.session_state.get("summary_text")
            if not summary_text:
                uploaded_name = st.session_state.get("uploaded_filename") or "uploaded.csv"
                summary_text = build_data_summary(df_chat, uploaded_name)
                st.session_state["summary_text"] = summary_text

            tmp_history = list(st.session_state.get("chat_history", []))
            tmp_history.append({"role": "user", "content": patched_user_text})

            with st.spinner("LLM が回答を生成しています..."):
                try:
                    content = call_llm_chat(summary_text, tmp_history, extra_system=extra_system)
                    code_block = ""
                    comment_block = content or ""
                    if content and "```python" in content:
                        before, _, rest = content.partition("```python")
                        code, _, after = rest.partition("```")
                        code_block = code.strip()
                        comment_block = (before + "\n\n" + after).strip()
                    if comment_block:
                        st.subheader("💬 LLMからの分析コメント")
                        st.markdown(comment_block)
                    if code_block:
                        st.subheader("📊 チャットで追加生成されたグラフ")
                        graph = {
                            "id": st.session_state.get("next_graph_id", 1),
                            "source": "chat",
                            "label": f"追加分析グラフ {st.session_state.get('next_graph_id', 1)}",
                            "code": code_block,
                        }
                        st.session_state.setdefault("graphs", []).append(graph)
                        st.session_state["next_graph_id"] = graph["id"] + 1
                        render_graph(graph, df_chat)
                    st.session_state.setdefault("chat_history", []).append({"role": "user", "content": patched_user_text})
                    st.session_state["chat_history"].append({"role": "assistant", "content": content})
                except Exception as e:
                    st.error(f"チャット分析中にエラーが発生しました: {e}")


_chat_section()

st.markdown("---")

# ── ベース分析（6項目）──
st.markdown("## 🔬 ベース分析（6項目）")
st.markdown(
    "重回帰分析・ABC分析・マーケットバスケット分析など、"
    "より深い視点から売上データを分析します。  \n"
    "データが不足する場合は **ダミーデータで分析イメージを表示** し、"
    "必要なデータを案内します。"
)

if st.session_state.get("show_additional"):
    df_for_analysis = st.session_state.get("df")
    if df_for_analysis is None:
        st.info(
            "DB からデータがまだ取得されていません。"
            "以下はすべてダミーデータによる分析イメージです。"
        )
        df_for_analysis = pd.DataFrame({
            "注文番号": [],
            "合計金額(税込)": [],
        })
    run_additional_analyses(df_for_analysis)
else:
    st.info("「🔄 DB からデータを取得」を実行するとベース分析が自動表示されます。")
